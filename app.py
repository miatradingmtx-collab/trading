# ==============================================================================
#                      MODERN TRADING WEB SERVICE (REST API)
# ==============================================================================
# NOTA IMPORTANTE SOBRE ARQUITECTURA:
# Este servicio web ha sido diseñado utilizando arquitectura REST moderna y formato JSON,
# reemplazando el formato XML/SOAP clásico.
#
# ¿Por qué los Web Services XML tradicionales (SOAP/WSDL) están obsoletos aquí?
# 1. TradingView y Notion no soportan XML nativamente para este tipo de flujos.
#    TradingView envía alertas en JSON, y la API de Notion consume estrictamente JSON.
# 2. XML es extremadamente pesado ("verboso") debido a las etiquetas de apertura y cierre.
#    JSON es ligero, rápido de transmitir y nativo en Python y JavaScript.
# 3. SOAP/XML requiere esquemas complejos (WSDL). REST/JSON utiliza FastAPI, que es el
#    estándar de la industria para microservicios de alto rendimiento y baja latencia.
# ==============================================================================

from fastapi import FastAPI, HTTPException, Request, BackgroundTasks, Header, Response
from fastapi.responses import HTMLResponse, StreamingResponse
from pydantic import BaseModel
import csv
from io import StringIO
from typing import Optional
import requests
import os
import datetime
import asyncio
from dotenv import load_dotenv
from mt5_executor_cloud import run_escaner_loop

# Cargar variables de entorno desde el archivo .env si existe localmente
load_dotenv()

import json
import firebase_admin
from firebase_admin import credentials, firestore
from google.cloud.firestore_v1.base_query import FieldFilter

# Inicializar Firebase de forma segura (redundancia local y nube)
firebase_inicializado = False
db = None

# Variables globales para caché del Dashboard
DASHBOARD_CACHE_DATA = None
DASHBOARD_CACHE_TIME = 0.0
import time

def invalidar_cache_dashboard():
    global DASHBOARD_CACHE_TIME, ULTIMO_FETCH_FIREBASE
    DASHBOARD_CACHE_TIME = 0.0
    ULTIMO_FETCH_FIREBASE = None

try:
    # 1. Intentar cargar desde un archivo local serviceAccountKey.json
    if os.path.exists("serviceAccountKey.json"):
        cred = credentials.Certificate("serviceAccountKey.json")
        firebase_admin.initialize_app(cred)
        firebase_inicializado = True
        db = firestore.client()
        print("| FIREBASE | Inicializado con éxito usando serviceAccountKey.json local.")
    
    # 2. Si no hay archivo, intentar cargar desde la variable de entorno JSON (para Render/Nube)
    elif os.getenv("FIREBASE_SERVICE_ACCOUNT_JSON"):
        raw_json = os.getenv("FIREBASE_SERVICE_ACCOUNT_JSON").strip()
        service_account_info = None
        try:
            service_account_info = json.loads(raw_json)
        except Exception as je:
            print(f"| FIREBASE | json.loads falló ({je}). Intentando método alternativo de extracción por Regex...")
            try:
                import re
                keys = [
                    "type", "project_id", "private_key_id", "private_key",
                    "client_email", "client_id", "auth_uri", "token_uri",
                    "auth_provider_x509_cert_url", "client_x509_cert_url", "universe_domain"
                ]
                extracted = {}
                for key in keys:
                    # Coincidir con "llave": "valor" con comillas simples o dobles
                    pattern = re.compile(
                        r'[\'"]' + re.escape(key) + r'[\'"]\s*:\s*[\'"](.*?)[\'"]',
                        re.DOTALL
                    )
                    match = pattern.search(raw_json)
                    if match:
                        val = match.group(1)
                        # Intentar descodificar con json.loads para resolver escapes estándar de JSON
                        try:
                            val = json.loads(f'"{val}"')
                        except Exception:
                            val = val.replace('\\"', '"').replace("\\'", "'")
                        
                        # Limpieza profunda de la clave privada
                        if key == "private_key":
                            # Convertir representaciones literales de saltos de línea en newlines reales
                            val = val.replace('\\\\n', '\n').replace('\\n', '\n')
                            # Resolver slashes escapados comunes en base64 (\/ -> /)
                            val = val.replace('\\/', '/')
                            # Eliminar cualquier diagonal invertida remanente para evitar fallos de PEM
                            val = val.replace('\\', '')
                        else:
                            val = val.replace('\\\\n', '\n').replace('\\n', '\n')
                        
                        extracted[key] = val
                
                if "private_key" in extracted and "client_email" in extracted:
                    service_account_info = extracted
                    print("| FIREBASE | Datos de cuenta de servicio extraídos con éxito vía Regex.")
                else:
                    raise ValueError("Faltan campos esenciales (private_key o client_email) tras extracción por Regex.")
            except Exception as e2:
                print(f"| FIREBASE ERROR | Falló también la extracción por Regex: {e2}")
                raise e2
        
        if service_account_info:
            cred = credentials.Certificate(service_account_info)
            firebase_admin.initialize_app(cred)
            firebase_inicializado = True
            db = firestore.client()
            print("| FIREBASE | Inicializado con éxito usando variable de entorno.")
    else:
        print("| FIREBASE WARNING | No se encontró archivo serviceAccountKey.json ni variable de entorno. Firebase no guardará datos.")
except Exception as e:
    print(f"| FIREBASE ERROR | Falló la inicialización de Firebase: {e}")

# ==============================================================================
# AUTO-INICIALIZADOR VIP DE ACTIVOS
# Si un activo nuevo llega via webhook o se detecta en el broker y NO existe
# en la matriz de Firebase, esta función lo crea automáticamente con el esquema
# completo del modelo de inteligencia financiera de Mia.
#
# OPTIMIZACIÓN DE TOKENS FIREBASE:
# - Usa el set ACTIVOS_INICIALIZADOS en RAM como primera barrera.
# - Si el activo ya está en el set → 0 lecturas a Firestore (costo $0).
# - Solo lee/escribe Firebase si el activo es genuinamente nuevo.
# ==============================================================================
def auto_inicializar_activo(activo: str) -> bool:
    """Inicializa un activo nuevo en la trading_matrix con el esquema completo de Mia si no existe."""
    global firebase_inicializado, db, ACTIVOS_INICIALIZADOS
    if not firebase_inicializado or db is None:
        return False
    activo_norm = normalizar_activo(activo)
    
    # 🔑 BARRERA DE RAM: Si ya está en caché, no gastamos ni un token de Firebase
    if activo_norm in ACTIVOS_INICIALIZADOS:
        return False
    
    try:
        doc_ref = db.collection("trading_matrix").document(activo_norm)
        doc = doc_ref.get()  # Solo se ejecuta si NO está en el caché RAM
        if doc.exists:
            # Ya existía en Firebase pero no estaba en RAM → agregar al caché
            ACTIVOS_INICIALIZADOS.add(activo_norm)
            return False
        
        # Es genuinamente nuevo: crear con el esquema completo de Mia
        precio_ref = 1.0
        if activo_norm in ["XAUUSD"]:
            precio_ref = 2300.0
        elif activo_norm in ["GBPJPY", "USDJPY", "EURJPY", "CHFJPY", "CADJPY", "AUDJPY", "NZDJPY"]:
            precio_ref = 170.0
        elif activo_norm in ["BTC", "ETH"]:
            precio_ref = 60000.0
        elif activo_norm in ["NAS100", "SPX500", "US30"]:
            precio_ref = 18000.0

        esquema_activo = {
            "activo": activo_norm,
            "estado_ejecucion": "INACTIVO",
            "ultimo_update": datetime.datetime.now(datetime.timezone.utc).isoformat(),
            "score_porcentaje": 0.0,
            "gatillo_entrada": False,
            "precio_referencia": precio_ref,
            "confirmaciones_tecnicas": {
                "soporte_resistencia_activo": False,
                "ema_50_200_crossover": False,
                "rsi_sobrecompra_sobreventa": False,
                "medias_moviles_alineadas": False,
                "poc_price": False,
                "smc_codes": []
            },
            "confirmaciones_fundamentales": {
                "noticias_impacto_favorables": False,
                "ipo_liquidez_positiva": False,
                "spo_liquidez_positiva": False
            },
            "confirmaciones_institucionales": {
                "dark_pools_amortizado": True,
                "dark_pools_url_valid": False,
                "whales_perdieron_fuerza": False,
                "heatmap_ordenes_limite": False
            },
            "aprendizaje_mia": {
                "modo_aprendiz_activo": True,
                "trades_totales": 0,
                "trades_ganados": 0,
                "win_rate_historico": 50.0,
                "racha_actual": 0,
                "sentimiento_alcista": False,
                "factor_ajuste_probabilidad": 0.0
            }
        }
        doc_ref.set(esquema_activo)
        ACTIVOS_INICIALIZADOS.add(activo_norm)  # Agregar al caché RAM inmediatamente
        print(f"| FIREBASE AUTO-VIP | ✔ Nuevo activo '{activo_norm}' matriculado automáticamente con esquema Mia completo.")
        return True
    except Exception as e:
        print(f"| FIREBASE AUTO-VIP ERROR | No se pudo inicializar '{activo}': {e}")
        return False

# Inicialización de la aplicación FastAPI (El estándar moderno de Web Services)
app = FastAPI(
    title="Trading Automation Bridge",
    description="Servidor puente moderno para conectar TradingView con Notion, Grok y Excel",
    version="1.0.0"
)

@app.on_event("startup")
async def startup_event():
    # Inicializar la base de datos de Firebase si está conectada
    global firebase_inicializado, db
    if firebase_inicializado and db is not None:
        try:
            # Lista de activos a validar
            # Lista VIP base (pares del broker)
            activos_vip = ["GBPJPY", "GBPUSD", "EURUSD", "XAUUSD", "AUDUSD", "NZDCAD",
                           "SPX500", "NAS100", "US30", "USDJPY", "USDCAD", "EURGBP",
                           "GBPCAD", "CHFJPY", "USDCHF", "EURJPY"]
            
            print("| FIREBASE VIP | Cargando activos ya existentes en Firebase (1 sola lectura batch)...")
            
            # 1 SOLA LECTURA BATCH: lee toda la colección de una vez
            docs_existentes = db.collection("trading_matrix").stream()
            for doc in docs_existentes:
                ACTIVOS_INICIALIZADOS.add(doc.id)  # Poblar caché RAM con los que ya existen
            
            print(f"| FIREBASE VIP | {len(ACTIVOS_INICIALIZADOS)} activos ya cargados en caché RAM: {sorted(ACTIVOS_INICIALIZADOS)}")
            
            # Solo inicializar los que NO estén ya en Firebase
            nuevos = [a for a in activos_vip if a not in ACTIVOS_INICIALIZADOS]
            if nuevos:
                print(f"| FIREBASE VIP | Inicializando {len(nuevos)} activos VIP nuevos: {nuevos}")
                for activo in nuevos:
                    auto_inicializar_activo(activo)
            else:
                print("| FIREBASE VIP | Todos los activos VIP ya están matriculados. Sin lecturas adicionales.")
            print("| FIREBASE VIP | ✔ Verificación de matriz VIP completada.")
        except Exception as e:
            print(f"| FIREBASE ERROR | Falló la auto-inicialización en startup: {e}")
            
    # Lanzar el escáner asíncrono de MetaAPI en segundo plano si está activado en el entorno
    if os.getenv("RUN_SCANNER_CLOUD", "false").lower() == "true":
        asyncio.create_task(run_escaner_loop())



# Configuración de variables de entorno (Coloca aquí tus llaves seguras)
NOTION_TOKEN = os.getenv("NOTION_TOKEN", "secret_TU_TOKEN_DE_NOTION")
NOTION_DATABASE_ID = os.getenv("NOTION_DATABASE_ID", "TU_DATABASE_ID_DE_NOTION")
GROK_API_KEY = os.getenv("GROK_API_KEY", "TU_LLAVE_DE_GROK")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "TU_LLAVE_DE_GEMINI")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "TU_LLAVE_DE_OPENAI")
BRIDGE_ACCESS_TOKEN = os.getenv("BRIDGE_ACCESS_TOKEN", "tu-token-seguro-de-acceso")

def verificar_token(authorization: Optional[str] = Header(None)):
    expected = f"Bearer {BRIDGE_ACCESS_TOKEN}"
    if not authorization or authorization != expected:
        raise HTTPException(status_code=401, detail="Token de acceso inválido o ausente")



# ------------------------------------------------------------------------------
# 1. MODELOS DE DATOS (Validación automática de la alerta de TradingView)
# ------------------------------------------------------------------------------
class TradeAlert(BaseModel):
    activo: str                # Ej: "EURUSD", "BTCUSD", "AAPL"
    accion: str                # Ej: "COMPRA", "VENTA"
    precio: float              # Ej: 1.0854, 68450.00
    stop_loss: Optional[float] = None
    take_profit: Optional[float] = None
    estrategia: str            # Ej: "RSI_Divergence", "MACD_Cross"
    pnl: Optional[float] = 0.0 # Beneficio/pérdida (para registrar cierres)
    ticket: Optional[str] = "" # Número de ticket/operación de MT5 (ahora como string por si MetaApi usa IDs)
    lotaje: Optional[float] = 0.01        # Volumen/Lotes de la operación
    temporalidad: Optional[str] = "1H"    # Temporalidad Swing (1H, 2H, 4H, 8H)
    es_crypto: Optional[bool] = False     # Indicador 24/7

class MarketAnomaly(BaseModel):
    activo: str                # Ej: "NASDAQ100", "SP500", "US30", "BTC"
    tipo: str                  # Ej: "DARK_POOL_PRINT", "OPTION_SWEEP", "BLOCK_TRADE"
    precio: float
    volumen_usd: float
    sentimiento: str           # Ej: "BULLISH", "BEARISH", "NEUTRAL"
    detalles: Optional[str] = None

class CollectiveMemoryRequest(BaseModel):
    memoria_compartida: str


# ------------------------------------------------------------------------------
# 2. FUNCIONES DE INTEGRACIÓN (Notion & Grok)
# ------------------------------------------------------------------------------
def enviar_a_notion(alert: TradeAlert):
    """Llamada a la API REST de Notion (JSON) para insertar el registro"""
    url = "https://api.notion.com/v1/pages"
    headers = {
        "Authorization": f"Bearer {NOTION_TOKEN}",
        "Content-Type": "application/json",
        "Notion-Version": "2022-06-28"
    }
    
    # Formato de datos JSON requerido por la API de Notion
    payload = {
        "parent": { "database_id": NOTION_DATABASE_ID },
        "properties": {
            "Activo": {
                "title": [
                    { "text": { "content": alert.activo } }
                ]
            },
            "Acción": {
                "select": { "name": alert.accion }
            },
            "Precio": {
                "number": alert.precio
            },
            "Stop Loss": {
                "number": alert.stop_loss if alert.stop_loss else 0.0
            },
            "Take Profit": {
                "number": alert.take_profit if alert.take_profit else 0.0
            },
            "Estrategia": {
                "rich_text": [
                    { "text": { "content": alert.estrategia } }
                ]
            },
            "PnL": {
                "number": alert.pnl if alert.pnl else 0.0
            },
            "Ticket": {
                "number": alert.ticket if alert.ticket else 0
            },
            "Fecha": {
                "date": { "start": datetime.datetime.now().isoformat() }
            }
        }
    }
    
    try:
        response = requests.post(url, headers=headers, json=payload)
        if response.status_code == 200:
            print(f"| NOTION | Entrada para {alert.activo} registrada con éxito.")
            return True
        else:
            print(f"| NOTION ERROR | Código {response.status_code}: {response.text}")
            return False
    except Exception as e:
        print(f"| NOTION EXCEPTION | Ocurrió un error al conectar: {e}")
        return False

def actualizar_excel_local(alert: TradeAlert):
    try:
        import openpyxl
        from openpyxl import load_workbook
        import datetime
        import os
        
        # Trabajar únicamente con la Bitacora de entradas 2025
        archivo = None
        rutas_posibles = [
            "Bitacora de entradas 2025.xlsx",
            "c:/Users/ecybe/OneDrive/Documentos/Trading/Bitacora de entradas 2025.xlsx"
        ]
        
        for ruta in rutas_posibles:
            if os.path.exists(ruta):
                archivo = ruta
                break
                
        if not archivo:
            print("| EXCEL WARNING | No se encontró la Bitacora de entradas 2025.xlsx. Omitiendo registro local.")
            return False
            
        wb = load_workbook(archivo)
        
        # 1. Asegurar la existencia de las hojas correctas
        sheet_names = wb.sheetnames
        ticket_sheet_name = None
        
        # Buscar si ya existe la hoja de tickets (incluso si tiene typos como ticek o tickets)
        for name in sheet_names:
            if name.lower() in ["ticket", "tickets", "ticek"]:
                ticket_sheet_name = name
                break
                
        if not ticket_sheet_name:
            if "Hoja2" in sheet_names:
                ws_hoja2 = wb["Hoja2"]
                ws_hoja2.title = "Ticket"
                ticket_sheet_name = "Ticket"
                print("| EXCEL | Renombrada 'Hoja2' a 'Ticket' en el libro.")
            else:
                wb.create_sheet("Ticket")
                ticket_sheet_name = "Ticket"
                print("| EXCEL | Creada nueva hoja 'Ticket' en el libro.")
                
        ws_main = wb["Hoja1"]
        ws_ticket = wb[ticket_sheet_name]
        
        # 2. Inicializar cabeceras de la hoja Ticket si está vacía
        ticket_headers = [cell.value for cell in ws_ticket[1]]
        if all(h is None for h in ticket_headers) or len(ticket_headers) == 0:
            headers_opt2 = ['COD', 'Año', 'Mes', 'Dia', 'Buy/Sell', 'Perdida', 'Ganada', '%', 'Activo', 'Temporalidad', 'Ganancia', 'RW', 'F', 'Hora']
            for col_idx, h in enumerate(headers_opt2, 1):
                ws_ticket.cell(row=1, column=col_idx, value=h)
            print("| EXCEL | Inicializada la hoja 'Ticket' con cabeceras estándar.")
            ticket_headers = headers_opt2
            
        # 3. Datos comunes de tiempo
        dias_semana = {0: "Lunes", 1: "Martes", 2: "Miercoles", 3: "Jueves", 4: "Viernes", 5: "Sabado", 6: "Domingo"}
        ahora = datetime.datetime.now()
        dia_str = dias_semana[ahora.weekday()]
        fecha_str = ahora.strftime("%Y-%m-%d")
        anio_short = ahora.year % 100
        mes_num = ahora.month
        hora_str = ahora.strftime("%H:%M:%S")
        
        accion_upper = alert.accion.upper()
        accion_normalizada = "Buy" if any(x in accion_upper for x in ["COMPRA", "BUY", "LONG", "B"]) else "Sell"
        
        pnl_val = alert.pnl if alert.pnl is not None else 0.0
        if pnl_val > 0.0:
            resultado_str = "Ganada"
        elif pnl_val < 0.0:
            resultado_str = "Perdida"
        else:
            resultado_str = "No se activo" if "CIERRE" not in accion_upper else "Be"
            
        # Función auxiliar de mapeo dinámico según encabezados
        def mapear_valores(headers_list):
            nueva_fila = [None] * len(headers_list)
            for idx, h_raw in enumerate(headers_list):
                if h_raw is None:
                    continue
                h = str(h_raw).strip().lower()
                
                if h in ["dia", "día"]:
                    nueva_fila[idx] = dia_str
                elif h == "cuenta":
                    nueva_fila[idx] = "Grafico"
                elif h in ["buy/sell", "action", "side", "compra/venta", "acción", "accion", "dirección", "direccion", "tipo"]:
                    nueva_fila[idx] = accion_normalizada
                elif h in ["entrada", "precio", "precio de entrada", "precio entrada", "entry", "precio_entrada"]:
                    nueva_fila[idx] = alert.precio
                elif h in ["sl", "stop loss", "stop_loss", "stop"]:
                    nueva_fila[idx] = alert.stop_loss
                elif h in ["tp", "take profit", "take_profit", "profit target"]:
                    nueva_fila[idx] = alert.take_profit
                elif h in ["lotaje", "lotes", "lot", "volume"]:
                    nueva_fila[idx] = alert.lotaje
                elif h in ["resultado", "status", "estado"]:
                    nueva_fila[idx] = resultado_str
                elif h in ["estado animico", "estado anímico"]:
                    nueva_fila[idx] = "Neutral"
                elif h in ["nombre", "activo", "ticker", "instrumento", "par", "symbol", "símbolo", "simbolo"]:
                    nueva_fila[idx] = alert.activo
                elif h in ["fecha", "date", "fecha de entrada", "fecha hora", "datetime", "fecha y hora"]:
                    nueva_fila[idx] = fecha_str
                elif h in ["monto", "ganancia", "ganancia usd", "pnl", "profit", "loss", "pérdida", "p&l"]:
                    nueva_fila[idx] = pnl_val
                elif h in ["temporalidad", "timeframe", "tf"]:
                    nueva_fila[idx] = alert.temporalidad
                elif h in ["comentarios", "estrategia", "strategy", "setup", "sistema", "nota", "notas"]:
                    nueva_fila[idx] = alert.estrategia
                elif h in ["ticket", "id", "orden", "operación", "operacion", "id_ticket", "ticket_id", "cod", "código", "codigo"]:
                    nueva_fila[idx] = alert.ticket
                elif h in ["año", "año short", "anio", "year"]:
                    nueva_fila[idx] = anio_short
                elif h == "mes":
                    nueva_fila[idx] = mes_num
                elif h == "hora":
                    nueva_fila[idx] = hora_str
            return nueva_fila

        # 4. Mapear y guardar en Hoja1
        headers_main = [str(cell.value).strip().lower() for cell in ws_main[1] if cell.value is not None]
        row_main = mapear_valores(headers_main)
        ws_main.append(row_main)
        
        # 5. Mapear y guardar en Ticket
        headers_ticket_processed = [str(cell.value).strip().lower() for cell in ws_ticket[1] if cell.value is not None]
        row_ticket = mapear_valores(headers_ticket_processed)
        ws_ticket.append(row_ticket)
        
        wb.save(archivo)
        print(f"| EXCEL SUCCESS | Operación guardada con éxito en Hoja1 y Ticket de {archivo}")
        return True
    except Exception as e:
        print(f"| EXCEL ERROR | No se pudo actualizar el archivo Excel: {e}")
        registrar_error_sistema("Excel Local", str(e))
        return False

def registrar_error_sistema(componente: str, mensaje: str):
    invalidar_cache_dashboard()
    """
    Registra errores críticos del sistema (Railway, Firebase, MetaAPI) en la colección mia_system_logs
    """
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        return
        
    try:
        import time
        doc_id = f"ERR_{int(time.time()*1000)}"
        db.collection("mia_system_logs").document(doc_id).set({
            "timestamp": datetime.datetime.now().isoformat(),
            "componente": componente,
            "mensaje": str(mensaje)
        })
    except:
        pass

def guardar_en_firestore(alert: TradeAlert, precio_yahoo: Optional[float] = None, precio_google: Optional[float] = None):
    """
    Registra la alerta de trading en la colección 'trading_alerts' de Firebase Firestore.
    """
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        print("| FIREBASE | Omitiendo registro en Firestore (Firebase no inicializado).")
        return False
        
    try:
        data = {
            "activo": alert.activo,
            "accion": alert.accion,
            "precio_alerta": alert.precio,
            "stop_loss": alert.stop_loss if alert.stop_loss else 0.0,
            "take_profit": alert.take_profit if alert.take_profit else 0.0,
            "estrategia": alert.estrategia,
            "pnl": alert.pnl if alert.pnl else 0.0,
            "ticket": alert.ticket if alert.ticket else 0,
            "precio_yahoo": precio_yahoo,
            "precio_google": precio_google,
            "timestamp": datetime.datetime.now()
        }
        
        # Guardar en la colección 'trading_alerts'
        # El método add genera un ID de documento aleatorio automáticamente
        doc_ref = db.collection("trading_alerts").add(data)
        print(f"| FIREBASE SUCCESS | Alerta guardada en Firestore. ID del documento: {doc_ref[1].id}")
        
        # Guardar en mia_audit_logs con el Ticket como ID (Para el Dashboard y KB)
        if alert.ticket:
            # PNL Accumulator: Si ya existe un registro previo de este ticket en Firestore (ej: CIERRE_PARCIAL),
            # recuperamos el PNL acumulado y lo sumamos para mostrar la ganancia real acumulada total.
            pnl_acumulado_previo = 0.0
            try:
                existente_doc = db.collection("mia_audit_logs").document(str(alert.ticket)).get()
                if existente_doc.exists:
                    exist_data = existente_doc.to_dict()
                    # Recuperar datos en caso de reporte tardío o incompleto
                    if alert.activo == "UNKNOWN":
                        alert.activo = exist_data.get("activo", "UNKNOWN")
                    if alert.precio == 0.0:
                        alert.precio = exist_data.get("precio_ejecucion", exist_data.get("precio", 0.0))
                    
                    # Almacenar PNL previo (de parciales anteriores)
                    pnl_acumulado_previo = float(exist_data.get("pnl", 0.0))
            except Exception as e:
                print(f"| FIREBASE | Error recuperando doc previo para {alert.ticket}: {e}")

            # Sumar el PNL actual al acumulado previo si es un cierre
            if alert.accion in ["CIERRE_TOTAL", "CIERRE_PARCIAL"]:
                alert.pnl = (alert.pnl if alert.pnl else 0.0) + pnl_acumulado_previo
            elif alert.pnl == 0.0 and pnl_acumulado_previo != 0.0:
                alert.pnl = pnl_acumulado_previo

            now_dt = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=-6)))
            utc_hour = datetime.datetime.utcnow().hour
            fecha_str = now_dt.strftime("%Y-%m-%d %H:%M:%S")
            iso_time = now_dt.isoformat()
            
            sesion = "NUEVA_YORK"
            h_local = now_dt.hour
            if 1 <= h_local < 6: sesion = "LONDRES"
            elif 6 <= h_local < 16: sesion = "NUEVA_YORK"
            else: sesion = "ASIA"
            activo_norm = normalizar_activo(alert.activo)
            score = 0
            poc_price = 0.0
            try:
                m_doc = db.collection("trading_matrix").document(activo_norm).get()
                if m_doc.exists:
                    m_data = m_doc.to_dict()
                    score = m_data.get("score_porcentaje", 0)
                    poc_price = m_data.get("confirmaciones_tecnicas", {}).get("poc_price", 0.0)
            except: pass
            
            motivo = "Rechazada por Matriz Técnica (Score bajo o Killzone)"
            if score >= 80:
                motivo = "En validación de riesgo por el Broker..."
                
            detalle_str = f"{alert.activo} | {fecha_str} | {sesion} | {alert.estrategia} | EVALUANDO SETUP | SCORE: {score}% | POC: {poc_price:.5f} | EJECUTADA EN MT5: NO | MOTIVO: {motivo}"
            
            # Determinar si es apertura de trade o cierre
            es_cierre = alert.accion in ["CIERRE_TOTAL", "CIERRE_PARCIAL"]
            
            if es_cierre:
                pnl_val = alert.pnl if alert.pnl else 0.0
                motivo_final = f"Cerrado en MT5 | PNL: ${pnl_val:.2f}" if alert.ticket else motivo
                ejecutada_flag = True  # El cierre confirma que el trade SI existio en MT5
                
                # RESETEAR SEMÁFORO A INACTIVO AL CERRAR LA POSICIÓN TOTALMENTE
                if alert.accion == "CIERRE_TOTAL":
                    try:
                        m_doc_ref = db.collection("trading_matrix").document(activo_norm)
                        m_doc_data = m_doc_ref.get().to_dict() or {}
                        m_doc_data["estado_ejecucion"] = "INACTIVO"
                        m_doc_ref.set(m_doc_data, merge=True)
                        print(f"| SEMÁFORO RESET | {activo_norm} reseteado a INACTIVO por CIERRE_TOTAL de ticket {alert.ticket}.")
                    except Exception as reset_e:
                        print(f"| SEMÁFORO RESET ERROR | No se pudo resetear estado para {activo_norm}: {reset_e}")
            else:
                # Es apertura COMPRA/VENTA
                motivo_final = "Ejecutada y Activa en Broker" if alert.ticket else motivo
                ejecutada_flag = True if alert.ticket else False
            
            audit_ref = db.collection("mia_audit_logs").document(str(alert.ticket))
            audit_data = {
                "ticket": str(alert.ticket),
                "activo": alert.activo,
                "accion": alert.accion,
                "estrategia": alert.estrategia,
                "pnl": alert.pnl if alert.pnl else 0.0,
                "ultima_actualizacion": iso_time,
                "timestamp": iso_time,
                "fecha": fecha_str,
                "score": score,
                "poc_price": poc_price,
                "ejecutada_mt5": ejecutada_flag,
                "motivo": motivo_final,
                "detalle_setup": detalle_str
            }
            # Usamos merge=True para no sobreescribir el precio y score si ya fue guardado por la apertura
            audit_ref.set(audit_data, merge=True)
            print(f"| AUDIT LOG SUCCESS | Ticket {alert.ticket} guardado/actualizado en mia_audit_logs.")
            
            # Actualizar Caché Global en RAM
            global GLOBAL_AUDIT_LOGS
            if GLOBAL_AUDIT_LOGS is not None:
                encontrado = False
                for i, log in enumerate(GLOBAL_AUDIT_LOGS):
                    if log.get("ticket") == str(alert.ticket):
                        GLOBAL_AUDIT_LOGS[i].update(audit_data)
                        encontrado = True
                        break
                if not encontrado:
                    GLOBAL_AUDIT_LOGS.append(audit_data)
                    
            invalidar_cache_dashboard()
            
        return True
    except Exception as e:
        print(f"| FIREBASE ERROR | Error al guardar en Firestore: {e}")
        registrar_error_sistema("Firebase (Alerta/Audit)", str(e))
        return False


BOTPRESS_WEBHOOK_URL = os.getenv("BOTPRESS_WEBHOOK_URL", "")
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "8914319073:AAHmF9BTxqgGG2XYn3whnXKe8RlJpzYG9Jk")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "")

def notificar_telegram(mensaje: str):
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        print("| TELEGRAM | Faltan credenciales, omitiendo mensaje.")
        return
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {
        "chat_id": TELEGRAM_CHAT_ID,
        "text": mensaje,
        "parse_mode": "Markdown"
    }
    try:
        response = requests.post(url, json=payload, timeout=5)
        if response.status_code == 200:
            print("| TELEGRAM | Notificación enviada con éxito.")
        else:
            print(f"| TELEGRAM | Error al enviar: {response.text}")
    except Exception as e:
        print(f"| TELEGRAM ERROR | {e}")
        registrar_error_sistema("Telegram", str(e))

def notificar_botpress_mia(activo: str, data: dict):
    # Enviar también a Telegram
    mensaje_tg = f"🤖 *MIA TRADING AI*\n\nNuevos datos para *{activo}*:\nEjecutando lógica en la nube."
    if "accion" in data:
        mensaje_tg = f"🤖 *MIA TRADING AI*\n\n🔥 *{data['accion']}* en *{activo}*\nPrecio: {data.get('precio', 'N/A')}"
    notificar_telegram(mensaje_tg)

    if not BOTPRESS_WEBHOOK_URL:
        print("| BOTPRESS | Webhook no configurado, omitiendo notificación a Mia.")
        return
    
    payload = {
        "activo": activo,
        "score": data.get("score_porcentaje", 0),
        "fundamental": data.get("confirmaciones_fundamentales", {}),
        "tecnico": data.get("confirmaciones_tecnicas", {})
    }
    try:
        requests.post(BOTPRESS_WEBHOOK_URL, json=payload, timeout=5)
        print(f"| BOTPRESS | Mia notificada exitosamente sobre setup en {activo}.")
    except Exception as e:
        print(f"| BOTPRESS ERROR | No se pudo notificar a Mia: {e}")
        registrar_error_sistema("Botpress", str(e))

def recalcular_score_ponderado(data: dict) -> float:
    score = 0.0
    tech = data.get("confirmaciones_tecnicas", {})
    
    # 1. Indicadores Macro (Filtros de Tendencia y Agotamiento)
    ma_alineada = tech.get("medias_moviles_alineadas", False)
    rsi_extremo = tech.get("rsi_sobrecompra_sobreventa", False) or tech.get("rsi_extremo", False)
    
    if not (ma_alineada or rsi_extremo):
        return 0.0  # Sin dirección clara ni zona de reversión, se rechaza
        
    if ma_alineada: score += 20
    if rsi_extremo: score += 20
    
    # 2. Confirmadores de Zonas Clave (POC y Soportes/Resistencias)
    if tech.get("soporte_resistencia_activo"): 
        score += 10 # (En mt5_executor_cloud, el POC ya se cuenta como soporte/resistencia)
        
    if tech.get("poc_price", 0.0) > 0:
        score += 10 # Bonus por tener confirmación clara de Perfil de Volumen
    
    # 3. Módulos SMC e ICT (Institucional)
    smc_codes = tech.get("smc_codes", [])
    
    # Pesos estructurales (Se suman a los indicadores para buscar >= 80%)
    if 1 in smc_codes: score += 40  # Order Block (OB)
    if 2 in smc_codes: score += 40  # Fair Value Gap (FVG)
    if 3 in smc_codes: score += 30  # Breaker Block
    if 4 in smc_codes: score += 20  # Liquidity Sweep
        
    # Firebase es el único juez de la validación. Permitimos scores > 100% para mostrar fuerza extrema.
    return score

def normalizar_activo(activo: str) -> str:
    """Mapea símbolos de trading comunes a los 8 activos clave de Firebase"""
    act = activo.upper().strip()
    if act in ["NASDAQ100", "NASDAQ", "NQ", "QQQ", "US100"]:
        return "NASDAQ100"
    if act in ["SP500", "SPY", "ES", "S&P500"]:
        return "SP500"
    if act in ["US30", "DJI", "YM", "DOW"]:
        return "US30"
    if act in ["BTC", "BTCUSD", "BITCOIN"]:
        return "BTC"
    if act in ["GBPJPY", "GBP-JPY"]:
        return "GBPJPY"
    if act in ["GBPUSD", "GBP-USD"]:
        return "GBPUSD"
    if act in ["EURUSD", "EUR-USD"]:
        return "EURUSD"
    if act in ["XAUUSD", "GOLD", "ORO", "GC"]:
        return "XAUUSD"
    return act

def procesar_anomalia_firestore(anomaly: MarketAnomaly):
    """
    Actualiza la matriz de trading en Firestore basándose en anomalías de Dark Pools u órdenes de bloque.
    """
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        print("| FIREBASE | Omitiendo procesamiento de anomalía (Firebase no inicializado).")
        return False
        
    try:
        activo_normalizado = normalizar_activo(anomaly.activo)
        doc_ref = db.collection("trading_matrix").document(activo_normalizado)
        doc = doc_ref.get()
        
        if not doc.exists:
            print(f"| FIREBASE ERROR | El activo '{activo_normalizado}' no está inicializado en la colección 'trading_matrix'.")
            return False
            
        data = doc.to_dict()
        is_bullish = anomaly.sentimiento.upper() == "BULLISH"
        
        if "confirmaciones_institucionales" not in data:
            data["confirmaciones_institucionales"] = {"dark_pools_compra_masiva": False, "heatmap_ordenes_limite": False}
            
        # Actualizar indicador según el tipo de anomalía
        if anomaly.tipo.upper() in ["DARK_POOL_PRINT", "BLOCK_TRADE"]:
            data["confirmaciones_institucionales"]["dark_pools_compra_masiva"] = is_bullish
            print(f"| FIREBASE | Actualizando Dark Pools de {activo_normalizado} a: {is_bullish}")
        elif anomaly.tipo.upper() == "HEATMAP_ORDER":
            data["confirmaciones_institucionales"]["heatmap_ordenes_limite"] = is_bullish
            print(f"| FIREBASE | Actualizando Heatmap de {activo_normalizado} a: {is_bullish}")
            
        # Calcular el Score Porcentaje total basado en el nuevo modelo Institucional (100 pts)
        score = recalcular_score_ponderado(data)
        data["score_porcentaje"] = round(score, 2)
        
        # El umbral configurado por el usuario es del 80% al 90%
        # Usamos 80% como umbral mínimo para activar el gatillo
        data["gatillo_entrada"] = score >= 80.0
        data["ultimo_update"] = datetime.datetime.now(datetime.timezone.utc).isoformat() if hasattr(datetime, "timezone") else datetime.datetime.now().isoformat()
        
        doc_ref.set(data)
        print(f"| FIREBASE SUCCESS | Matriz de {activo_normalizado} actualizada. Score: {data['score_porcentaje']}% | Gatillo: {data['gatillo_entrada']}")
        return True
    except Exception as e:
        print(f"| FIREBASE ERROR | Error al procesar anomalía en Firestore: {e}")
        registrar_error_sistema("Firebase (Anomalía)", str(e))
        return False

def obtener_precio_yahoo(activo: str) -> Optional[float]:
    """
    Obtiene el precio en tiempo real directamente desde Yahoo Finance.
    Esto permite validar o enriquecer la alerta recibida de TradingView.
    """
    try:
        import yfinance as yf
        ticker_nombre = activo
        
        # Correcciones comunes de formato para Yahoo Finance:
        # Forex: EURUSD -> EURUSD=X
        if len(activo) == 6 and activo.isupper() and not activo.endswith("=X"):
            # Si parece Forex tradicional (ej: EURUSD, GBPUSD)
            if any(pair in activo for pair in ["EUR", "GBP", "USD", "JPY", "AUD", "CAD", "CHF"]):
                ticker_nombre = f"{activo}=X"
        
        # Crypto: BTCUSD -> BTC-USD
        elif activo.startswith("BTC") and len(activo) == 6:
            ticker_nombre = "BTC-USD"
            
        ticker = yf.Ticker(ticker_nombre)
        hist = ticker.history(period="1d", interval="1m")
        if not hist.empty:
            precio_actual = hist['Close'].iloc[-1]
            print(f"| YAHOO FINANCE | Precio obtenido para {ticker_nombre}: {precio_actual}")
            return float(precio_actual)
        
        # Intento de respaldo si el intervalo de 1m falla
        hist_diario = ticker.history(period="1d")
        if not hist_diario.empty:
            precio_actual = hist_diario['Close'].iloc[-1]
            print(f"| YAHOO FINANCE | Precio (diario) para {ticker_nombre}: {precio_actual}")
            return float(precio_actual)
            
        print(f"| YAHOO FINANCE | No hay datos históricos para {ticker_nombre}")
        return None
    except Exception as e:
        print(f"| YAHOO FINANCE ERROR | Ocurrió un error al obtener precio: {e}")
        registrar_error_sistema("Yahoo Finance", str(e))
        return None

def obtener_precio_google(activo: str) -> Optional[float]:
    """
    Obtiene el precio en tiempo real raspando Google Finance como respaldo a Yahoo Finance.
    Esto proporciona redundancia de grado institucional.
    """
    try:
        from bs4 import BeautifulSoup
        
        # Formatear ticker para Google Finance. Ej: AAPL -> AAPL:NASDAQ
        # Para Forex: EURUSD -> EUR-USD
        ticker_nombre = activo
        if len(activo) == 6 and activo.isupper():
            if any(pair in activo for pair in ["EUR", "GBP", "USD", "JPY", "AUD", "CAD", "CHF"]):
                ticker_nombre = f"{activo[:3]}-{activo[3:]}" # EUR-USD
                
        url = f"https://www.google.com/finance/quote/{ticker_nombre}"
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        }
        
        response = requests.get(url, headers=headers, timeout=5)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            # Google Finance almacena el precio principal en un div con clase "YMl7ec"
            price_div = soup.find("div", class_="YMl7ec")
            if price_div:
                # Limpiar símbolos monetarios
                price_str = price_div.text.replace("$", "").replace("€", "").replace("£", "").replace(",", "").strip()
                precio_actual = float(price_str)
                print(f"| GOOGLE FINANCE | Precio obtenido para {ticker_nombre}: {precio_actual}")
                return precio_actual
                
        print(f"| GOOGLE FINANCE | No se pudo extraer precio de la página para {ticker_nombre}")
        return None
    except Exception as e:
        print(f"| GOOGLE FINANCE ERROR | Ocurrió un error al raspar precio: {e}")
        registrar_error_sistema("Google Finance", str(e))
        return None





# ------------------------------------------------------------------------------
# 3. ENDPOINTS DEL SERVICIO WEB (Ruta que escucha a TradingView)
# ------------------------------------------------------------------------------
@app.get("/")
def ruta_principal():
    return {
        "estado": "activo",
        "servicio": "Trading Automation Bridge",
        "arquitectura": "REST API (JSON)",
        "nota": "Para enviar alertas, usa el método POST en /webhook"
    }

# ------------------------------------------------------------------------------
# KEEP-ALIVE & HEALTH CHECK (para GitHub Actions, n8n y UptimeRobot)
# ------------------------------------------------------------------------------
@app.get("/health")
def health_check():
    """
    Endpoint de salud para keep-alive.
    Usado por:
    - GitHub Actions cron job (cada 14 min)
    - n8n workflow (cada 14 min)
    - UptimeRobot (si se configura)
    """
    return {
        "status": "ok",
        "service": "Mia Trading Bot",
        "timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "firebase": "connected" if firebase_inicializado else "disconnected",
        "uptime": "24/7"
    }

# ------------------------------------------------------------------------------
# WEBHOOK MARKET ALERT (n8n despierta a Mia cuando detecta keywords de mercado)
# ------------------------------------------------------------------------------
class MarketAlertPayload(BaseModel):
    source: str                          # "n8n-monitor"
    alert_type: str                      # "market_keyword_detected"
    keywords: Optional[str] = ""        # "FOMC, NFP, GOLD"
    summary: Optional[str] = ""         # Resumen del alert
    timestamp: Optional[str] = ""       # ISO timestamp

@app.post("/webhook_market_alert")
async def webhook_market_alert(
    payload: MarketAlertPayload,
    request: Request,
    authorization: Optional[str] = Header(None)
):
    """
    Recibe alertas de n8n cuando detecta palabras clave críticas en páginas de mercado.
    Despierta a Mia y registra el evento en Firebase.
    """
    # Validar token de acceso
    token = ACCESS_TOKEN
    if authorization and authorization.startswith("Bearer "):
        provided = authorization.split(" ")[1]
        if provided != token and token not in ["tu-token-seguro-de-acceso", "", None]:
            raise HTTPException(status_code=401, detail="Token de acceso inválido")

    print(f"| N8N ALERT | Alerta de mercado recibida: {payload.alert_type}")
    print(f"| N8N ALERT | Keywords: {payload.keywords}")
    print(f"| N8N ALERT | Resumen: {payload.summary}")

    # Guardar en Firebase si está disponible
    if firebase_inicializado and db is not None:
        try:
            db.collection("market_alerts").add({
                "source": payload.source,
                "alert_type": payload.alert_type,
                "keywords": payload.keywords,
                "summary": payload.summary,
                "timestamp": payload.timestamp or datetime.datetime.now(datetime.timezone.utc).isoformat(),
                "procesado": True
            })
            print("| FIREBASE | Alerta de mercado guardada en Firestore.")
        except Exception as e:
            print(f"| FIREBASE ERROR | No se pudo guardar alerta de mercado: {e}")
            registrar_error_sistema("Firebase (Market Alert)", str(e))

    # Consultar Mia (Gemini/Grok) con el contexto del mercado si hay keywords críticos
    mia_response = None
    if payload.keywords and len(payload.keywords) > 10:
        try:
            # Crear un TradeAlert simulado para usar las funciones existentes de análisis
            fake_alert = TradeAlert(
                activo="XAUUSD",
                accion="ANÁLISIS",
                precio=0.0,
                stop_loss=0.0,
                take_profit=0.0,
                estrategia=f"N8N Monitor: {payload.keywords}"
            )
            if GEMINI_API_KEY and GEMINI_API_KEY not in ["TU_LLAVE_DE_GEMINI", ""]:
                mia_response = consultar_analisis_gemini(fake_alert)
            elif GROK_API_KEY and GROK_API_KEY not in ["TU_LLAVE_DE_GROK", ""]:
                mia_response = consultar_analisis_grok(fake_alert)
        except Exception as e:
            print(f"| MIA ERROR | No se pudo obtener análisis de Mia: {e}")
            registrar_error_sistema("Mia AI (Analysis)", str(e))

    return {
        "status": "received",
        "alert_type": payload.alert_type,
        "keywords_detected": payload.keywords,
        "mia_analysis": mia_response or "Mia no disponible (configura GEMINI_API_KEY o GROK_API_KEY)",
        "timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat()
    }

def recalcular_memoria_colectiva():
    """Recalcula el resumen ejecutivo de mia_kb y lo guarda en system_memory/mia_collective"""
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        return

    try:
        ind_docs = db.collection("mia_kb").document("indicadores_impacto").collection("detalle").stream()
        mejor_indicador = {"nombre": "ninguno", "win_rate": 0.0}
        peor_indicador = {"nombre": "ninguno", "win_rate": 100.0}
        
        for idoc in ind_docs:
            idata = idoc.to_dict()
            if idata.get("trades_con_indicador", 0) > 0:
                wr = idata.get("win_rate_indicador", 0.0)
                if wr > mejor_indicador["win_rate"]:
                    mejor_indicador = {"nombre": idoc.id, "win_rate": wr}
                if wr < peor_indicador["win_rate"]:
                    peor_indicador = {"nombre": idoc.id, "win_rate": wr}
                    
        ses_docs = db.collection("mia_kb").document("sesiones_rendimiento").collection("detalle").stream()
        mejor_sesion = {"nombre": "ninguna", "win_rate": 0.0}
        for sdoc in ses_docs:
            sdata = sdoc.to_dict()
            if sdata.get("trades_totales", 0) > 0 and sdata.get("win_rate", 0.0) > mejor_sesion["win_rate"]:
                mejor_sesion = {"nombre": sdoc.id, "win_rate": sdata["win_rate"]}
                
        pat_docs = db.collection("mia_kb").document("patrones_ict_smc").collection("detalle").stream()
        patron_estrella = {"nombre": "ninguno", "win_rate": 0.0}
        for pdoc in pat_docs:
            pdata = pdoc.to_dict()
            if pdata.get("ocurrencias", 0) > 0 and pdata.get("win_rate", 0.0) > patron_estrella["win_rate"]:
                patron_estrella = {"nombre": pdoc.id, "win_rate": pdata["win_rate"]}

        resumen = {
            "mejor_indicador": mejor_indicador["nombre"],
            "mejor_indicador_win_rate": mejor_indicador["win_rate"],
            "peor_indicador": peor_indicador["nombre"],
            "peor_indicador_win_rate": peor_indicador["win_rate"],
            "mejor_sesion": mejor_sesion["nombre"],
            "mejor_sesion_win_rate": mejor_sesion["win_rate"],
            "patron_estrella_ict_smc": patron_estrella["nombre"],
            "patron_estrella_win_rate": patron_estrella["win_rate"],
            "modo_escucha": True,
            "ultimo_calculo": datetime.datetime.now(datetime.timezone.utc).isoformat() if hasattr(datetime, "timezone") else datetime.datetime.now().isoformat(),
            "resumen_operativo": (
                f"Mia KB v1.0 | Mejor indicador: {mejor_indicador['nombre']} ({mejor_indicador['win_rate']}% WR) | "
                f"Mejor sesion: {mejor_sesion['nombre']} ({mejor_sesion['win_rate']}% WR) | "
                f"Patron estrella: {patron_estrella['nombre']} ({patron_estrella['win_rate']}% WR) | "
                f"Peor indicador: {peor_indicador['nombre']} ({peor_indicador['win_rate']}% WR)"
            )
        }
        db.collection("system_memory").document("mia_collective").set(resumen)
        print(f"| KB MIA | Memoria colectiva recalculada exitosamente.")
    except Exception as e:
        print(f"| KB MIA ERROR | Error recalculando memoria colectiva: {e}")
        registrar_error_sistema("Mia KB (Collective)", str(e))

def determinar_tipo_salida_ticket(ticket: str):
    """
    Stored Procedure de análisis: Determina el tipo exacto de salida de un ticket
    agrupando todos los logs históricos asociados en Firebase.
    """
    if not ticket or str(ticket) == "0" or str(ticket) == "None":
        return "DESCONOCIDO"
        
    global GLOBAL_AUDIT_LOGS
    t_logs = []
    if GLOBAL_AUDIT_LOGS:
        t_logs = [l for l in GLOBAL_AUDIT_LOGS if str(l.get("ticket")) == str(ticket)]
        
    if not t_logs:
        return "DESCONOCIDO"
        
    acciones = [str(l.get("accion", "")).upper() for l in t_logs]
    pnls = [float(l.get("pnl", 0.0)) for l in t_logs]
    comentarios = [str(l.get("estrategia", "")).upper() for l in t_logs]
    
    has_cierre_total = "CIERRE_TOTAL" in acciones
    has_cierre_parcial = "CIERRE_PARCIAL" in acciones or any("PARCIAL" in c for c in comentarios)
    
    if not has_cierre_total:
        return "ABIERTO"
        
    if has_cierre_parcial:
        cierre_total_log = next((l for l in t_logs if str(l.get("accion")).upper() == "CIERRE_TOTAL"), None)
        if cierre_total_log:
            pnl_cierre = float(cierre_total_log.get("pnl", 0.0))
            if abs(pnl_cierre) <= 1.5:
                return "PARCIAL_BE"
            else:
                return "PARCIAL_MANUAL"
        else:
            return "PARCIAL_BE"
    else:
        cierre_total_log = next((l for l in t_logs if str(l.get("accion")).upper() == "CIERRE_TOTAL"), None)
        if cierre_total_log:
            ct_comentario = str(cierre_total_log.get("estrategia", "")).upper()
            pnl_cierre = float(cierre_total_log.get("pnl", 0.0))
            
            if pnl_cierre < 0:
                return "SL_ORIGINAL"
            elif "DESAPARICION" in ct_comentario or "MANUAL" in ct_comentario:
                return "MANUAL_DIRECTO"
            else:
                return "TP_COMPLETO"
        else:
            pnl_final = sum(pnls)
            if pnl_final < 0:
                return "SL_ORIGINAL"
            else:
                return "TP_COMPLETO"

def actualizar_aprendizaje_mia(activo: str, pnl: float, ticket: str = ""):
    """
    Stored Procedure (SP): Actualiza la base de conocimiento (mia_kb) leyendo las 
    confirmaciones de trading_matrix, calculando sesiones y patrones ICT/SMC.
    """
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        return
        
    try:
        activo_normalizado = normalizar_activo(activo)
        doc_ref = db.collection("trading_matrix").document(activo_normalizado)
        doc = doc_ref.get()
        
        if not doc.exists:
            return
            
        data = doc.to_dict()
        es_ganado = pnl > 0.0
        
        # 1. Snapshot de confirmaciones
        confirmaciones_activas = []
        for cat in ["confirmaciones_tecnicas", "confirmaciones_fundamentales", "confirmaciones_institucionales"]:
            if cat in data:
                for campo, valor in data[cat].items():
                    if valor is True:
                        confirmaciones_activas.append(campo)
                        
        # 2. Detectar sesion
        hora_utc = datetime.datetime.now(datetime.timezone.utc).hour
        if 0 <= hora_utc < 8:
            sesion = "asia"
        elif 8 <= hora_utc < 13:
            sesion = "london"
        elif 13 <= hora_utc < 15:
            sesion = "overlap_london_ny"
        elif 15 <= hora_utc < 21:
            sesion = "new_york"
        else:
            sesion = "asia"
            
        # 3. Actualizar Indicadores de Impacto
        for indicador in confirmaciones_activas:
            try:
                ind_ref = db.collection("mia_kb").document("indicadores_impacto").collection("detalle").document(indicador)
                ind_doc = ind_ref.get()
                if ind_doc.exists:
                    ind_data = ind_doc.to_dict()
                else:
                    ind_data = {"trades_con_indicador": 0, "trades_ganados_con": 0, "trades_perdidos_con": 0, "pnl_acumulado": 0.0, "win_rate_indicador": 0.0}
                    
                ind_data["trades_con_indicador"] = ind_data.get("trades_con_indicador", 0) + 1
                if es_ganado:
                    ind_data["trades_ganados_con"] = ind_data.get("trades_ganados_con", 0) + 1
                else:
                    ind_data["trades_perdidos_con"] = ind_data.get("trades_perdidos_con", 0) + 1
                    
                ind_data["pnl_acumulado"] = ind_data.get("pnl_acumulado", 0.0) + pnl
                if ind_data["trades_con_indicador"] > 0:
                    ind_data["win_rate_indicador"] = round(
                        (ind_data["trades_ganados_con"] / ind_data["trades_con_indicador"]) * 100, 2
                    )
                ind_data["ultima_actualizacion"] = datetime.datetime.now(datetime.timezone.utc).isoformat() if hasattr(datetime, "timezone") else datetime.datetime.now().isoformat()
                ind_ref.set(ind_data)
            except Exception as e:
                print(f"| KB MIA WARN | Error actualizando indicador {indicador}: {e}")
                registrar_error_sistema("Mia KB (Indicador)", str(e))
                
        # 4. Actualizar Sesion de Rendimiento
        try:
            ses_ref = db.collection("mia_kb").document("sesiones_rendimiento").collection("detalle").document(sesion)
            ses_doc = ses_ref.get()
            if ses_doc.exists:
                ses_data = ses_doc.to_dict()
            else:
                ses_data = {"trades_totales": 0, "trades_ganados": 0, "pnl_total": 0.0, "win_rate": 0.0}
                
            ses_data["trades_totales"] = ses_data.get("trades_totales", 0) + 1
            if es_ganado:
                ses_data["trades_ganados"] = ses_data.get("trades_ganados", 0) + 1
            ses_data["pnl_total"] = ses_data.get("pnl_total", 0.0) + pnl
            if ses_data["trades_totales"] > 0:
                ses_data["win_rate"] = round(
                    (ses_data["trades_ganados"] / ses_data["trades_totales"]) * 100, 2
                )
            ses_data["ultima_actualizacion"] = datetime.datetime.now(datetime.timezone.utc).isoformat() if hasattr(datetime, "timezone") else datetime.datetime.now().isoformat()
            ses_ref.set(ses_data)
        except Exception as e:
            print(f"| KB MIA WARN | Error actualizando sesion {sesion}: {e}")
            registrar_error_sistema("Mia KB (Sesión)", str(e))
            
        # 5. Detectar y Actualizar Patrones ICT/SMC
        ict_fields = {
            "order_block_detectado": "OB",
            "fvg_detectado": "FVG",
            "breaker_block_detectado": "BRK",
            "sweep_liquidez_detectado": "SWEEP",
            "soporte_resistencia_activo": "SR"
        }
        patron_key_parts = sorted([ict_fields[f] for f in confirmaciones_activas if f in ict_fields])
        if patron_key_parts:
            patron_key = "_".join(patron_key_parts)
            try:
                pat_ref = db.collection("mia_kb").document("patrones_ict_smc").collection("detalle").document(patron_key)
                pat_doc = pat_ref.get()
                if pat_doc.exists:
                    pat_data = pat_doc.to_dict()
                else:
                    pat_data = {
                        "combo": patron_key_parts,
                        "metodologia": "ICT/SMC",
                        "ocurrencias": 0,
                        "ganados": 0,
                        "perdidos": 0,
                        "win_rate": 0.0,
                        "pnl_generado": 0.0,
                        "tickets_ganadores": [],
                        "tickets_perdedores": [],
                        "cierres_tp_completo": 0,
                        "cierres_sl_original": 0,
                        "cierres_parcial_be": 0,
                        "cierres_parcial_manual": 0,
                        "cierres_manual_directo": 0
                    }
                    
                pat_data["ocurrencias"] = pat_data.get("ocurrencias", 0) + 1
                
                # Clasificar tipo de cierre para el patrón
                tipo_salida = determinar_tipo_salida_ticket(ticket)
                pat_data["cierres_tp_completo"] = pat_data.get("cierres_tp_completo", 0)
                pat_data["cierres_sl_original"] = pat_data.get("cierres_sl_original", 0)
                pat_data["cierres_parcial_be"] = pat_data.get("cierres_parcial_be", 0)
                pat_data["cierres_parcial_manual"] = pat_data.get("cierres_parcial_manual", 0)
                pat_data["cierres_manual_directo"] = pat_data.get("cierres_manual_directo", 0)
                
                if tipo_salida == "TP_COMPLETO":
                    pat_data["cierres_tp_completo"] += 1
                elif tipo_salida == "SL_ORIGINAL":
                    pat_data["cierres_sl_original"] += 1
                elif tipo_salida == "PARCIAL_BE":
                    pat_data["cierres_parcial_be"] += 1
                elif tipo_salida == "PARCIAL_MANUAL":
                    pat_data["cierres_parcial_manual"] += 1
                elif tipo_salida == "MANUAL_DIRECTO":
                    pat_data["cierres_manual_directo"] += 1
                
                if es_ganado:
                    pat_data["ganados"] = pat_data.get("ganados", 0) + 1
                    if ticket:
                        ganadores_arr = pat_data.get("tickets_ganadores", [])
                        if str(ticket) not in ganadores_arr:
                            ganadores_arr.append(str(ticket))
                            pat_data["tickets_ganadores"] = ganadores_arr
                else:
                    pat_data["perdidos"] = pat_data.get("perdidos", 0) + 1
                    if ticket:
                        perdidos_arr = pat_data.get("tickets_perdedores", [])
                        if str(ticket) not in perdidos_arr:
                            perdidos_arr.append(str(ticket))
                            pat_data["tickets_perdedores"] = perdidos_arr
                        
                pat_data["pnl_generado"] = pat_data.get("pnl_generado", 0.0) + pnl
                if pat_data["ocurrencias"] > 0:
                    pat_data["win_rate"] = round((pat_data["ganados"] / pat_data["ocurrencias"]) * 100, 2)
                pat_data["ultima_actualizacion"] = datetime.datetime.now(datetime.timezone.utc).isoformat() if hasattr(datetime, "timezone") else datetime.datetime.now().isoformat()
                pat_ref.set(pat_data)
            except Exception as e:
                print(f"| KB MIA WARN | Error actualizando patron ICT/SMC {patron_key}: {e}")
                registrar_error_sistema("Mia KB (Patrón)", str(e))
                
        # 6. Recalcular Memoria Colectiva
        recalcular_memoria_colectiva()
        
        # Opcional: Actualizar la estadística legacy si existe
        if "aprendizaje_mia" in data:
            apoyo = data["aprendizaje_mia"]
            apoyo["trades_totales"] = apoyo.get("trades_totales", 0) + 1
            if es_ganado:
                apoyo["trades_ganados"] = apoyo.get("trades_ganados", 0) + 1
                apoyo["racha_actual"] = max(1, apoyo.get("racha_actual", 0) + 1)
            else:
                apoyo["racha_actual"] = min(-1, apoyo.get("racha_actual", 0) - 1)
            apoyo["win_rate_historico"] = round((apoyo["trades_ganados"] / apoyo["trades_totales"]) * 100.0, 2)
            data["aprendizaje_mia"] = apoyo
            doc_ref.set(data)
            
        print(f"| KB MIA | Aprendizaje registrado para {activo_normalizado}. PnL: {pnl} | Sesion: {sesion}")
    except Exception as e:
        print(f"| APRENDIZAJE MIA ERROR | Error al procesar aprendizaje de trade: {e}")
        registrar_error_sistema("Aprendizaje MIA", str(e))

@app.post("/webhook")
def recibir_alerta(alert: TradeAlert, background_tasks: BackgroundTasks):
    """
    Ruta que recibe el Webhook de TradingView en formato JSON.
    Usa BackgroundTasks para procesar la API de Notion y Grok en segundo plano,
    permitiendo que TradingView reciba una respuesta instantánea (baja latencia).
    """
    # RECUPERACIÓN DE DATOS ANTES DE PROCESAR:
    # Si viene como UNKNOWN desde mt5_executor_cloud, recuperamos de Firestore inmediatamente
    if alert.activo == "UNKNOWN" and alert.ticket:
        try:
            doc = db.collection("mia_audit_logs").document(str(alert.ticket)).get()
            if doc.exists:
                exist_data = doc.to_dict()
                alert.activo = exist_data.get("activo", "UNKNOWN")
                if alert.pnl == 0.0:
                    alert.pnl = exist_data.get("pnl", 0.0)
                if alert.precio == 0.0:
                    alert.precio = exist_data.get("precio_ejecucion", exist_data.get("precio", 0.0))
        except Exception as e:
            pass

    print(f"\n========================================================")
    print(f"ALERTA RECIBIDA DE TRADINGVIEW: {alert.accion} en {alert.activo}")
    print(f"Precio Alerta: {alert.precio} | Estrategia: {alert.estrategia}")
    print(f"========================================================")
    
    # AUTO-VIP: Si el activo no está en la matriz, lo registramos automáticamente con el esquema completo
    if alert.activo and alert.activo != "UNKNOWN":
        auto_inicializar_activo(alert.activo)
    
    # 0. Lógica de Horarios (Forex cerrado en fin de semana, Crypto 24/7)
    es_cripto_activo = alert.es_crypto or alert.activo.startswith("BTC") or alert.activo.startswith("ETH") or "USD" not in alert.activo and alert.activo != "XAUUSD"
    ahora = datetime.datetime.now(datetime.timezone.utc)
    if not es_cripto_activo:
        # Viernes después de 21:00 UTC hasta Domingo a las 21:00 UTC es fin de semana en Forex (aprox)
        if ahora.weekday() == 5 or (ahora.weekday() == 4 and ahora.hour >= 21) or (ahora.weekday() == 6 and ahora.hour < 21):
            print(f"| REGLA DE HORARIO | Mercado Forex cerrado. Rechazando orden de {alert.activo}.")
            return {"resultado": "rechazado", "mensaje": "Mercado Forex cerrado en fin de semana."}

    # 0.5 Filtro de Killzones por Activo (Usando hora NY / EST)
    # Convertimos UTC a EST (restando 5 horas o 4 en Daylight Saving, usaremos aprox UTC-4 para verano, UTC-5 invierno. Simplificando a UTC-4)
    hora_ny = (ahora.hour - 4) % 24
    
    # Definición de Killzones
    en_asia = (20 <= hora_ny <= 23) or (0 <= hora_ny < 2) # 20:00 a 02:00
    en_londres = (2 <= hora_ny < 6) # 02:00 a 06:00
    en_ny = (7 <= hora_ny < 11) # 07:00 a 11:00
    en_killzone_activa = False
    
    activo_upper = alert.activo.upper()
    if es_cripto_activo:
        en_killzone_activa = True # Crypto 24/7
    elif "EUR" in activo_upper or "USD" in activo_upper or "XAU" in activo_upper:
        if en_londres or en_ny: en_killzone_activa = True
    elif "JPY" in activo_upper or "AUD" in activo_upper or "NZD" in activo_upper:
        if en_asia or en_londres: en_killzone_activa = True
        
    if not en_killzone_activa and alert.accion in ["COMPRA", "VENTA"]:
        print(f"| KILLZONE | Trade rechazado para {alert.activo}. Fuera de sus ventanas de alta liquidez (Hora NY actual: {hora_ny}:00).")
        return {"resultado": "rechazado", "mensaje": "Fuera de Killzone de liquidez."}

    # 1. Obtener precios de validación de ambas fuentes (Yahoo y Google)
    precio_yahoo = obtener_precio_yahoo(alert.activo)
    precio_google = obtener_precio_google(alert.activo)
    
    # Imprimir validaciones cruzadas en el servidor
    print(f"| VALIDACIÓN | TradingView: {alert.precio} | Yahoo: {precio_yahoo} | Google: {precio_google}")
    
    # 2. Lógica de enriquecimiento con redundancia inteligente
    if alert.precio == 0.0:
        if precio_yahoo:
            alert.precio = precio_yahoo
            print(f"| ENRIQUECIMIENTO | Precio establecido mediante Yahoo Finance: {alert.precio}")
        elif precio_google:
            alert.precio = precio_google
            print(f"| ENRIQUECIMIENTO | Fallback exitoso: Precio establecido mediante Google Finance: {alert.precio}")
        else:
            print("| ENRIQUECIMIENTO ADVERTENCIA | No se pudo obtener cotización de ninguna fuente externa.")

    # 3. Ejecutar el guardado en Notion en segundo plano
    background_tasks.add_task(enviar_a_notion, alert)
    background_tasks.add_task(actualizar_excel_local, alert)
    background_tasks.add_task(guardar_en_firestore, alert, precio_yahoo, precio_google)
    
    # 4. Modo Aprendiz (KB de Mia): Sincronizar resultados si es un cierre con PnL
    if alert.pnl != 0.0 or "CIERRE" in alert.accion.upper():
        background_tasks.add_task(actualizar_aprendizaje_mia, alert.activo, alert.pnl, alert.ticket)
        
    # 5. Notificar a Mia (Botpress) de que hubo un movimiento (Apertura o Cierre)
    if BOTPRESS_WEBHOOK_URL:
        payload_mia = {
            "evento": "trade_ejecutado",
            "activo": alert.activo,
            "accion": alert.accion,
            "precio": alert.precio,
            "estrategia": alert.estrategia
        }
        def avisar_mia():
            try:
                requests.post(BOTPRESS_WEBHOOK_URL, json=payload_mia, timeout=5)
                print(f"| BOTPRESS | Mia notificada de {alert.accion} en {alert.activo}.")
            except Exception as e:
                print(f"| BOTPRESS ERROR | No se pudo despertar a Mia: {e}")
                registrar_error_sistema("Botpress", str(e))
        background_tasks.add_task(avisar_mia)
    
    return {
        "resultado": "recibido",
        "mensaje": f"Procesando operación de {alert.accion} para {alert.activo}",
        "precio_utilizado": alert.precio,
        "precio_yahoo": precio_yahoo,
        "precio_google": precio_google,
        "timestamp": datetime.datetime.now().isoformat()
    }

@app.get("/webhook_get")
def recibir_alerta_get(
    activo: str,
    accion: str,
    precio: float = 0.0,
    estrategia: str = "manual_get",
    pnl: float = 0.0,
    background_tasks: BackgroundTasks = BackgroundTasks()
):
    """
    Ruta alternativa GET para pruebas rápidas de texto directamente desde el navegador web.
    Ejemplo de uso: http://localhost:8000/webhook_get?activo=BTCUSD&accion=COMPRA&precio=68000
    
    LIMITACIÓN CRÍTICA DE GET: Las imágenes NO se pueden enviar por aquí debido a las restricciones 
    de longitud de caracteres en la URL (~2048 caracteres). Para imágenes o archivos binarios pesados, 
    el método POST es obligatorio.
    """
    alert = TradeAlert(
        activo=activo,
        accion=accion,
        precio=precio,
        estrategia=estrategia,
        pnl=pnl
    )
    
    precio_yahoo = obtener_precio_yahoo(alert.activo)
    precio_google = obtener_precio_google(alert.activo)
    
    if alert.precio == 0.0:
        if precio_yahoo:
            alert.precio = precio_yahoo
        elif precio_google:
            alert.precio = precio_google

    background_tasks.add_task(enviar_a_notion, alert)
    background_tasks.add_task(actualizar_excel_local, alert)
    background_tasks.add_task(guardar_en_firestore, alert, precio_yahoo, precio_google)

    # 4. Modo Aprendiz (KB de Mia): Sincronizar resultados si es un cierre con PnL
    if alert.pnl != 0.0 or "CIERRE" in alert.accion.upper():
        background_tasks.add_task(actualizar_aprendizaje_mia, alert.activo, alert.pnl)
    
    if BOTPRESS_WEBHOOK_URL:
        payload_mia = {
            "evento": "trade_ejecutado",
            "activo": alert.activo,
            "accion": alert.accion,
            "precio": alert.precio,
            "estrategia": alert.estrategia
        }
        def avisar_mia_get():
            try:
                requests.post(BOTPRESS_WEBHOOK_URL, json=payload_mia, timeout=5)
                print(f"| BOTPRESS GET | Mia notificada de {alert.accion} en {alert.activo}.")
            except Exception as e:
                print(f"| BOTPRESS ERROR | No se pudo despertar a Mia: {e}")
                registrar_error_sistema("Botpress (GET)", str(e))
        background_tasks.add_task(avisar_mia_get)
        
    # <-- AÑADIDO: Notificar también a Telegram -->
    mensaje_tg = f"🤖 *MIA TRADING AI*\n\n🔥 *{alert.accion}* en *{alert.activo}*\nPrecio: {alert.precio}"
    notificar_telegram(mensaje_tg)
    
    return {
        "resultado": "recibido_via_get",
        "mensaje": f"Procesando operación de {alert.accion} para {alert.activo}",
        "precio_utilizado": alert.precio,
        "precio_yahoo": precio_yahoo,
        "precio_google": precio_google,
        "timestamp": datetime.datetime.now().isoformat()
    }


@app.get("/test_buy")
async def test_buy(simbolo: str = "XAUUSD", lote: float = 0.01):
    """
    Ruta de prueba para abrir una posición de compra en el broker de forma inmediata.
    Ejemplo de uso: http://localhost:8080/test_buy?simbolo=XAUUSD&lote=0.01
    """
    from mt5_executor_cloud import abrir_posicion_test
    res = await abrir_posicion_test(simbolo, lote)
    return {"status": "success", "result": res}


# BLOQUEADO TEMPORALMENTE (Evitar ejecuciones externas por bots de ping)
# @app.get("/test_boolean")
async def test_boolean(activo: str = "XAUUSD", lote: float = 0.01):
    """
    Ruta de prueba para validar la lógica booleana en Firebase:
    1. Fuerza las 11 confirmaciones a True en Firestore para el activo.
    2. Lee el documento de Firestore y cuenta cuántas confirmaciones están en True.
    3. Si la cantidad de confirmaciones True es >= 8, ejecuta una compra de prueba en MetaAPI.
    4. Restaura las confirmaciones originales del activo.
    """
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        return {"status": "error", "message": "Firebase no inicializado"}
        
    activo_normalizado = normalizar_activo(activo)
    doc_ref = db.collection("trading_matrix").document(activo_normalizado)
    
    try:
        # Guardar estado original
        doc = doc_ref.get()
        original_data = doc.to_dict() if doc.exists else None
        
        # 1. Forzar las 11 confirmaciones a True para la prueba
        test_data = {
            "confirmaciones_tecnicas": {
                "soporte_resistencia_activo": True,
                "medias_moviles_alineadas": True,
                "rsi_sobrecompra_sobreventa": True,
                "order_block_detectado": True,
                "fvg_detectado": True,
                "breaker_block_detectado": True,
                "sweep_liquidez_detectado": True
            },
            "confirmaciones_fundamentales": {
                "noticias_impacto_favorables": True,
                "ipo_liquidez_positiva": True,
                "spo_liquidez_positiva": True
            },
            "confirmaciones_institucionales": {
                "dark_pools_compra_masiva": True,
                "heatmap_ordenes_limite": True
            },
            "score_porcentaje": 100.0,
            "activo": activo_normalizado
        }
        doc_ref.set(test_data, merge=True)
        print(f"| TEST BOOLEAN | Confirmaciones forzadas a True para {activo_normalizado}")
        
        # 2. Leer de nuevo y contar
        doc_test = doc_ref.get()
        data_test = doc_test.to_dict()
        
        # Contar confirmaciones True
        true_count = 0
        categories = ["confirmaciones_tecnicas", "confirmaciones_fundamentales", "confirmaciones_institucionales"]
        for cat in categories:
            if cat in data_test:
                for field, val in data_test[cat].items():
                    if val is True or val == 1:
                        true_count += 1
                        
        print(f"| TEST BOOLEAN | Conteo de confirmaciones True en Firebase para {activo_normalizado}: {true_count}")
        
        result_msg = ""
        # 3. Validar si es >= 8
        if true_count >= 8:
            print(f"| TEST BOOLEAN SUCCESS | Conteo ({true_count}) >= 8. Autorizando compra de prueba...")
            from mt5_executor_cloud import abrir_posicion_test
            trade_res = await abrir_posicion_test(activo_normalizado, lote)
            result_msg = f"Aprobado (Conteo: {true_count} >= 8). Trade result: {trade_res}"
        else:
            result_msg = f"Rechazado (Conteo: {true_count} < 8)."
            
        # 4. Restaurar original si existía
        if original_data:
            doc_ref.set(original_data)
            print(f"| TEST BOOLEAN | Estado original restaurado para {activo_normalizado}")
            
        return {
            "status": "success",
            "activo": activo_normalizado,
            "confirmaciones_true_detectadas": true_count,
            "resultado_validacion": result_msg
        }
        
    except Exception as e:
        print(f"| TEST BOOLEAN ERROR | Ocurrió un error al conectar: {e}")
        registrar_error_sistema("Test Boolean", str(e))
        return {"status": "error", "message": str(e)}



@app.post("/webhook_anomaly")
def recibir_anomalia(anomaly: MarketAnomaly, background_tasks: BackgroundTasks):
    """
    Ruta para recibir anomalías de flujo institucional (Dark Pools / Opciones / Heatmap)
    de proveedores de datos (Unusual Whales / Tradytics) vía n8n.
    """
    print(f"\n========================================================")
    print(f"ANOMALÍA DETECTADA: {anomaly.tipo} en {anomaly.activo}")
    print(f"Volumen: ${anomaly.volumen_usd:,.2f} | Sentimiento: {anomaly.sentimiento}")
    print(f"========================================================")
    
    # Validar el umbral (solo procesamos anomalías institucionales mayores a $5,000,000)
    # Puedes ajustar este umbral según tus preferencias de volumen
    UMBRAL_MINIMO_USD = 5000000.0
    if anomaly.volumen_usd < UBRAL_MINIMO_USD:
        print(f"| FILTRO | Anomalía ignorada. Volumen (${anomaly.volumen_usd:,.2f}) menor al umbral mínimo (${UMBRAL_MINIMO_USD:,.2f})")
        return {"resultado": "ignorado", "motivo": "volumen por debajo del umbral"}
        
    background_tasks.add_task(procesar_anomalia_firestore, anomaly)
    
    return {
        "resultado": "recibido",
        "mensaje": f"Procesando anomalía {anomaly.tipo} para {anomaly.activo} en segundo plano",
        "timestamp": datetime.datetime.now().isoformat()
    }


# ------------------------------------------------------------------------------
# NUEVOS WEBHOOKS PARA METATRADER 5 (INTEGRACIÓN CON EL EXECUTOR LOCAL)
# ------------------------------------------------------------------------------

@app.get("/get_matrix_activos")
def get_matrix_activos(authorization: Optional[str] = Header(None)):
    """
    Ruta para que n8n obtenga la lista de activos actualmente configurados en la matriz.
    Así n8n solo hace polling fundamental de los activos relevantes.
    """
    verificar_token(authorization)
    
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        raise HTTPException(status_code=503, detail="Firebase no inicializado")
        
    try:
        docs = db.collection("trading_matrix").stream()
        activos = [doc.id for doc in docs]
        return {"status": "success", "activos": activos}
    except Exception as e:
        print(f"| CLOUD ERROR | Error en get_matrix_activos: {e}")
        registrar_error_sistema("Cloud API (Get Matrix)", str(e))
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/get_asset_matrix")
def get_asset_matrix(activo: str, authorization: Optional[str] = Header(None)):
    """
    Ruta para obtener la matriz actual de confirmaciones de un activo específico.
    Utilizada por el executor local para validar liquidez institucional.
    """
    verificar_token(authorization)
    
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        raise HTTPException(status_code=503, detail="Firebase no inicializado")
        
    try:
        activo_normalizado = normalizar_activo(activo)
        doc_ref = db.collection("trading_matrix").document(activo_normalizado)
        doc = doc_ref.get()
        
        if not doc.exists:
            raise HTTPException(status_code=404, detail=f"Activo {activo_normalizado} no encontrado")
            
        return doc.to_dict()
    except HTTPException:
        raise
    except Exception as e:
        print(f"| CLOUD ERROR | Error al obtener matriz de activo {activo}: {e}")
        registrar_error_sistema("Cloud API (Get Asset)", str(e))
        raise HTTPException(status_code=500, detail=str(e))

class TechnicalUpdate(BaseModel):
    activo: str
    confirmaciones_tecnicas: dict
    killzone_activa: Optional[bool] = True

class FundamentalUpdate(BaseModel):
    activo: str
    noticias_impacto_favorables: Optional[bool] = None
    ipo_liquidez_positiva: Optional[bool] = None
    spo_liquidez_positiva: Optional[bool] = None

class MT5SetupRequest(BaseModel):
    activo: str
    accion: str
    precio: float
    estrategia: str

class SystemErrorLog(BaseModel):
    componente: str
    mensaje: str

@app.post("/webhook_log_error")
def api_webhook_log_error(err: SystemErrorLog, authorization: Optional[str] = Header(None)):
    verificar_token(authorization)
    registrar_error_sistema(err.componente, err.mensaje)
    return {"status": "ok"}

@app.post("/webhook_technical_update")
def webhook_technical_update(update: TechnicalUpdate, authorization: Optional[str] = Header(None)):
    """
    Ruta que recibe las confirmaciones técnicas en tiempo real calculadas por el script
    de MetaTrader 5 y actualiza la matriz en Firebase.
    """
    verificar_token(authorization)
    invalidar_cache_dashboard()
    
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        raise HTTPException(status_code=503, detail="Firebase no inicializado")
        
    try:
        activo_normalizado = normalizar_activo(update.activo)
        doc_ref = db.collection("trading_matrix").document(activo_normalizado)
        doc = doc_ref.get()
        
        if not doc.exists:
            raise HTTPException(status_code=404, detail=f"El activo {activo_normalizado} no existe en la matriz")
            
        data = doc.to_dict()
        
        if "confirmaciones_tecnicas" not in data:
            data["confirmaciones_tecnicas"] = {}
            
        # Actualizar confirmaciones técnicas
        for k, v in update.confirmaciones_tecnicas.items():
            if k == "smc_codes":
                data["confirmaciones_tecnicas"][k] = v
            else:
                data["confirmaciones_tecnicas"][k] = bool(v)
                
        # Limpiar booleanos legacy si existen en la base de datos
        legacy_keys = ["order_block_detectado", "fvg_detectado", "breaker_block_detectado", "sweep_liquidez_detectado"]
        for lk in legacy_keys:
            if lk in data["confirmaciones_tecnicas"]:
                del data["confirmaciones_tecnicas"][lk]
            
        # Calcular el Score Porcentaje total basado en el nuevo modelo Institucional (100 pts)
        score = recalcular_score_ponderado(data)
        data["score_porcentaje"] = round(score, 2)
        data["gatillo_entrada"] = score >= 80.0
        
        if data["gatillo_entrada"] and data.get("estado_ejecucion", "INACTIVO") == "INACTIVO":
            data["estado_ejecucion"] = "PENDIENTE_EJECUCIÓN"
            print(f"| SEMÁFORO | {activo_normalizado} ha cambiado a PENDIENTE_EJECUCIÓN")
        elif not data["gatillo_entrada"] and data.get("estado_ejecucion") != "INACTIVO":
            # Resetear semáforo si se perdió el setup
            data["estado_ejecucion"] = "INACTIVO"
            print(f"| SEMÁFORO | {activo_normalizado} ha cambiado a INACTIVO (Score insuficiente)")
            
        data["ultimo_update"] = datetime.datetime.now(datetime.timezone.utc).isoformat() if hasattr(datetime, "timezone") else datetime.datetime.now().isoformat()
        
        doc_ref.set(data)
        print(f"| FIREBASE SUCCESS | Confirmaciones técnicas de {activo_normalizado} actualizadas. Score: {data['score_porcentaje']}%")
        
        # --- GENERAR LOG DE EVALUACIÓN PARA EL LIVE FEED ---
        now_dt = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=-6)))
        fecha_str = now_dt.strftime("%Y-%m-%d %H:%M:%S")
        iso_time = now_dt.isoformat()
        
        utc_hour = datetime.datetime.utcnow().hour
        sesion = "NY"
        if 0 <= utc_hour < 7: sesion = "ASIA"
        elif 7 <= utc_hour < 12: sesion = "LONDRES"
        
        if score < 80:
            motivo = "Evaluación Continua (Score insuficiente"
            if update.killzone_activa is False:
                motivo += " y Fuera de Killzone)"
            else:
                motivo += ")"
        else:
            if update.killzone_activa is False:
                motivo = "Rechazada por Killzone (Fuera de horario)"
            else:
                motivo = "Setup Detectado (Esperando ejecución)"
                
        confs = []
        for k, v in update.confirmaciones_tecnicas.items():
            if k == "smc_codes" and isinstance(v, list) and v:
                confs.append("SMC")
            elif isinstance(v, bool) and v:
                confs.append(k.replace("_", " ").upper())
                
        confirmaciones_str = " + ".join(confs) if confs else "Setup Base"
        detalle_str = f"{activo_normalizado} | {fecha_str} | {sesion} | Escáner Cloud | {confirmaciones_str} | SCORE: {score}% | EJECUTADA EN MT5: NO | MOTIVO: {motivo}"
        
        import time
        eval_id = f"EVAL_{activo_normalizado}_{int(time.time())}"
        audit_ref = db.collection("mia_audit_logs").document(eval_id)
        audit_ref.set({
            "ticket": eval_id,
            "activo": activo_normalizado,
            "estrategia": "Escáner Cloud",
            "score": score,
            "ejecutada_mt5": False,
            "motivo": motivo,
            "fecha": fecha_str,
            "timestamp": iso_time,
            "detalle_setup": detalle_str,
            "confirmaciones_tecnicas": data.get("confirmaciones_tecnicas", {}),
            "confirmaciones_fundamentales": data.get("confirmaciones_fundamentales", {}),
            "confirmaciones_institucionales": data.get("confirmaciones_institucionales", {})
        })
        
        return {
            "status": "success",
            "activo": activo_normalizado,
            "score_porcentaje": data["score_porcentaje"],
            "gatillo_entrada": data["gatillo_entrada"]
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"| CLOUD ERROR | Error en webhook_technical_update: {e}")
        registrar_error_sistema("Webhook Scanner Cloud", str(e))
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/webhook_mt5_setup")
def webhook_mt5_setup(req: MT5SetupRequest, background_tasks: BackgroundTasks, authorization: Optional[str] = Header(None)):
    """
    Ruta que evalúa si el score del activo es >= 80% en Firebase, consulta a las IAs
    para el contexto fundamental/sentimiento geopolítico, y retorna la autorización final del trade.
    """
    verificar_token(authorization)
    invalidar_cache_dashboard()
    
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        raise HTTPException(status_code=503, detail="Firebase no inicializado")
        
    try:
        activo_normalizado = normalizar_activo(req.activo)
        doc_ref = db.collection("trading_matrix").document(activo_normalizado)
        doc = doc_ref.get()
        
        if not doc.exists:
            return {
                "authorized": False,
                "reason": f"Activo {activo_normalizado} no inicializado en Firestore"
            }
            
        data = doc.to_dict()
        
        # 1. VALIDACIÓN DE 'LIVE KEYS' (API Tokens de Notion, Firebase, y al menos una IA en .env)
        live_keys_notion = NOTION_TOKEN and NOTION_TOKEN != "secret_TU_TOKEN_DE_NOTION" and "coloca_aqui" not in NOTION_TOKEN
        live_keys_firebase = firebase_inicializado and db is not None
        live_keys_ai = (
            (GEMINI_API_KEY and GEMINI_API_KEY != "TU_LLAVE_DE_GEMINI" and "coloca_aqui" not in GEMINI_API_KEY) or
            (OPENAI_API_KEY and OPENAI_API_KEY != "TU_LLAVE_DE_OPENAI" and "coloca_aqui" not in OPENAI_API_KEY) or
            (GROK_API_KEY and GROK_API_KEY != "TU_LLAVE_DE_GROK" and "coloca_aqui" not in GROK_API_KEY)
        )
        # Descomentar la siguiente línea para habilitar Notion e IA como requisitos obligatorios (quitando el bypass)
        # live_keys_valid = live_keys_notion and live_keys_firebase and live_keys_ai
        
        # BYPASS ACTIVO: Solo Firebase es requerido para ejecutar en MT5
        live_keys_valid = live_keys_firebase
        
        if not live_keys_valid:
            detalles_faltantes = []
            if not live_keys_notion: detalles_faltantes.append("Notion API Token (Advertencia: No se registrará en Notion, pero la ejecución continuará si las demás APIs están bien)")
            if not live_keys_firebase: detalles_faltantes.append("Conexión Firestore de Firebase")
            if not live_keys_ai: detalles_faltantes.append("Al menos una API Key de IA (Gemini, ChatGPT o Grok)")
            
            return {
                "authorized": False,
                "reason": f"Fallo de validación de 'live keys' (APIs). Faltan/Inválidas: {', '.join(detalles_faltantes)}",
                "live_keys_valid": False
            }

        # 1.5 VALIDACIÓN DE SEMÁFORO DE EJECUCIÓN
        estado_actual = data.get("estado_ejecucion", "INACTIVO")
        if estado_actual != "PENDIENTE_EJECUCIÓN":
            return {
                "authorized": False,
                "reason": f"Semáforo no autorizado. El estado actual es '{estado_actual}', se requiere 'PENDIENTE_EJECUCIÓN' (Score >= 80%).",
                "estado_ejecucion": estado_actual
            }

        # 2. (REMOVIDO) VALIDACIÓN DE 'LEVEL KEYS'
        # Anteriormente se exigía Soporte/Resistencia, OB o BB de forma estricta.
        # Esto fue removido porque la metodología SMC ya valida estas estructuras
        # (incluyendo FVG y Sweep) y las pondera en el score. Si el score llega al 80%,
        # la estructura es matemáticamente válida según la configuración de Mia.
        # 3. VALIDACIÓN FINAL DE PROBABILIDAD ESTADÍSTICA (Score >= 80%)
        # El score debe ser mayor o igual al 80% como primera condición
        score = data.get("score_porcentaje", 0.0)
        gatillo = data.get("gatillo_entrada", False)
        score_valido = gatillo or (score >= 80.0)
        
        if not score_valido:
            return {
                "authorized": False,
                "reason": f"El score de validación ({score}%) es menor al 80% requerido.",
                "score_porcentaje": score
            }
            
        # Obtener memoria colectiva de Firestore si existe
        memoria_colectiva = None
        if firebase_inicializado and db is not None:
            try:
                mem_doc = db.collection("system_memory").document("mia_collective").get()
                if mem_doc.exists:
                    memoria_colectiva = mem_doc.to_dict().get("memoria_compartida")
                    print(f"| APRENDIZAJE MIA | Memoria colectiva cruzada cargada exitosamente.")
            except Exception as e:
                print(f"| APRENDIZAJE MIA ERROR | No se pudo leer la memoria colectiva cruzada: {e}")

        # Consultar IAs para el contexto geopolítico y fundamental
        alert = TradeAlert(
            activo=req.activo,
            accion=req.accion,
            precio=req.precio,
            estrategia=req.estrategia
        )
        
        analisis_ia = "No se pudo obtener análisis de ninguna IA."
        if GEMINI_API_KEY and GEMINI_API_KEY != "TU_LLAVE_DE_GEMINI":
            print("| IA | Consultando análisis a Google Gemini...")
            analisis_ia = consultar_analisis_gemini(alert, memoria_colectiva)
        elif OPENAI_API_KEY and OPENAI_API_KEY != "TU_LLAVE_DE_OPENAI":
            print("| IA | Consultando análisis a OpenAI ChatGPT...")
            analisis_ia = consultar_analisis_chatgpt(alert, memoria_colectiva)
        elif GROK_API_KEY and GROK_API_KEY != "TU_LLAVE_DE_GROK":
            print("| IA | Consultando análisis a xAI Grok...")
            analisis_ia = consultar_analisis_grok(alert, memoria_colectiva)
        else:
            print("| IA WARNING | Ninguna API Key de IA configurada. Usando fallback de análisis local.")
            analisis_ia = f"Filtro fundamental local aprobado por Mia. Memoria colectiva: {memoria_colectiva if memoria_colectiva else 'Ninguna'}. Operar con gestión de riesgo estricta."
            
        # Calcular SL y TP inteligentes basados en el activo
        precio_ej = req.precio
        tipo_orden = req.accion.upper()
        
        # Configuración por defecto dinámica (para Forex u otros si no están en la lista)
        if precio_ej < 5.0: # Pares Forex estándar (AUDUSD, NZDCAD, EURUSD...)
            pips_def = 0.0020
            sl = precio_ej - pips_def if tipo_orden == "COMPRA" else precio_ej + pips_def
            tp = precio_ej + (pips_def * 2.0) if tipo_orden == "COMPRA" else precio_ej - (pips_def * 2.0)
            lote = 0.1
        elif precio_ej < 300.0: # Pares JPY
            pips_def = 0.30
            sl = precio_ej - pips_def if tipo_orden == "COMPRA" else precio_ej + pips_def
            tp = precio_ej + (pips_def * 2.0) if tipo_orden == "COMPRA" else precio_ej - (pips_def * 2.0)
            lote = 0.1
        else: # Cripto, Índices o Oro
            sl = precio_ej - 200.0 if tipo_orden == "COMPRA" else precio_ej + 200.0
            tp = precio_ej + 400.0 if tipo_orden == "COMPRA" else precio_ej - 400.0
            lote = 0.1
        
        # Ajustes institucionales por tipo de activo
        if activo_normalizado in ["EURUSD", "GBPUSD"]:
            pips = 0.0020 if activo_normalizado == "EURUSD" else 0.0025
            sl = precio_ej - pips if tipo_orden == "COMPRA" else precio_ej + pips
            tp = precio_ej + (pips * 2.0) if tipo_orden == "COMPRA" else precio_ej - (pips * 2.0)
            lote = 0.5
        elif activo_normalizado in ["AUDUSD", "NZDCAD"]:
            # AUDUSD y NZDCAD requieren SL más holgado por spreads cruzados
            pips = 0.0035
            sl = precio_ej - pips if tipo_orden == "COMPRA" else precio_ej + pips
            tp = precio_ej + (pips * 2.0) if tipo_orden == "COMPRA" else precio_ej - (pips * 2.0)
            lote = 0.4
        elif activo_normalizado == "GBPJPY":
            pips = 0.35 # Subimos a 35 pips para darle holgura y evitar barridas rápidas
            sl = precio_ej - pips if tipo_orden == "COMPRA" else precio_ej + pips
            tp = precio_ej + (pips * 2.0) if tipo_orden == "COMPRA" else precio_ej - (pips * 2.0)
            lote = 0.3
        elif activo_normalizado == "XAUUSD":
            sl = precio_ej - 5.0 if tipo_orden == "COMPRA" else precio_ej + 5.0
            tp = precio_ej + 12.0 if tipo_orden == "COMPRA" else precio_ej - 12.0
            lote = 0.1
        elif activo_normalizado == "BTC":
            sl = precio_ej - 500.0 if tipo_orden == "COMPRA" else precio_ej + 500.0
            tp = precio_ej + 1500.0 if tipo_orden == "COMPRA" else precio_ej - 1500.0
            lote = 0.05
        elif activo_normalizado in ["NASDAQ100", "SP500", "US30"]:
            pct_sl = 0.01 if activo_normalizado != "SP500" else 0.007
            sl = precio_ej * (1.0 - pct_sl) if tipo_orden == "COMPRA" else precio_ej * (1.0 + pct_sl)
            tp = precio_ej * (1.0 + pct_sl * 2.5) if tipo_orden == "COMPRA" else precio_ej * (1.0 - pct_sl * 2.5)
            lote = 0.2
            
        sl = round(sl, 5)
        tp = round(tp, 5)
        probabilidad = score # Asignamos la probabilidad para evitar UnboundLocalError
        
        # Logs en segundo plano (Notion, Excel Local y Firestore)
        background_tasks.add_task(enviar_a_notion, alert)
        background_tasks.add_task(actualizar_excel_local, alert)
        background_tasks.add_task(guardar_en_firestore, alert, None, None)
        
        # Cambiar el semáforo a EJECUTADO
        data["estado_ejecucion"] = "EJECUTADO"
        doc_ref.set(data)
        
        print(f"| DECISIÓN CLOUD | Trade AUTORIZADO para {activo_normalizado}. Score: {score}%. Probabilidad: {probabilidad}%. SL: {sl} | TP: {tp}")
        
        return {
            "authorized": True,
            "activo": activo_normalizado,
            "accion": req.accion,
            "precio": req.precio,
            "lote": lote,
            "stop_loss": sl,
            "take_profit": tp,
            "analisis_ia": analisis_ia,
            "score_porcentaje": score,
            "probabilidad_exito": probabilidad
        }
    except Exception as e:
        print(f"| CLOUD ERROR | Error en webhook_mt5_setup: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/mia_trading_feed.xml")
def get_mia_trading_feed(authorization: Optional[str] = Header(None)):
    """
    Exposes Mia's trading learnings and sentiment as an RSS/XML feed.
    This feed can be consumed by n8n or other Mia instances to synchronize collective intelligence.
    """
    verificar_token(authorization)
    
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        raise HTTPException(status_code=503, detail="Firebase no inicializado")
        
    try:
        docs = db.collection("trading_matrix").stream()
        
        xml_items = []
        for doc in docs:
            activo_id = doc.id
            data = doc.to_dict()
            apoyo = data.get("aprendizaje_mia", {})
            
            raw_sent = apoyo.get('sentimiento_acumulado', 'NEUTRAL').upper()
            if raw_sent == "BULLISH":
                sentimiento_val = "True"
            elif raw_sent == "BEARISH":
                sentimiento_val = "False"
            else:
                sentimiento_val = "NEUTRAL"
            
            # Format as XML item
            item_xml = f"""
        <activo name="{activo_id}">
            <trades_totales>{apoyo.get('trades_totales', 0)}</trades_totales>
            <trades_ganados>{apoyo.get('trades_ganados', 0)}</trades_ganados>
            <win_rate_historico>{apoyo.get('win_rate_historico', 50.0)}</win_rate_historico>
            <racha_actual>{apoyo.get('racha_actual', 0)}</racha_actual>
            <sentimiento_acumulado>{sentimiento_val}</sentimiento_acumulado>
            <factor_ajuste_probabilidad>{apoyo.get('factor_ajuste_probabilidad', 0.0)}</factor_ajuste_probabilidad>
            <ultimo_update>{data.get('ultimo_update', '')}</ultimo_update>
            <score_porcentaje>{data.get('score_porcentaje', 0.0)}</score_porcentaje>
        </activo>"""
            xml_items.append(item_xml)
            
        xml_content = f"""<?xml version="1.0" encoding="UTF-8"?>
<mia_trading_learnings>
    <canal>
        <titulo>MIA Trading Bot learnings</titulo>
        <descripcion>Base de conocimiento (KB) de Mia en Trading Algoritmico</descripcion>
        <generacion>{datetime.datetime.now(datetime.timezone.utc).isoformat() if hasattr(datetime, 'timezone') else datetime.datetime.now().isoformat()}</generacion>
        <activos>{"".join(xml_items)}
        </activos>
    </canal>
</mia_trading_learnings>"""
        
        return Response(content=xml_content, media_type="application/xml")
    except Exception as e:
        print(f"| FEED XML ERROR | {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/update_collective_memory")
def update_collective_memory(req: CollectiveMemoryRequest, authorization: Optional[str] = Header(None)):
    """
    Allows n8n or another agent instance to update the cross-project collective memory of Mia.
    This string is injected into Mia's system prompts.
    """
    verificar_token(authorization)
    
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        raise HTTPException(status_code=503, detail="Firebase no inicializado")
        
    try:
        doc_ref = db.collection("system_memory").document("mia_collective")
        doc_ref.set({
            "memoria_compartida": req.memoria_compartida,
            "ultimo_update": datetime.datetime.now(datetime.timezone.utc).isoformat() if hasattr(datetime, "timezone") else datetime.datetime.now().isoformat()
        })
        print(f"| FIREBASE SUCCESS | Memoria colectiva de MIA actualizada con éxito.")
        return {"status": "success", "message": "Memoria colectiva actualizada con éxito"}
    except Exception as e:
        print(f"| FIREBASE ERROR | Error al actualizar memoria colectiva: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/webhook_fundamental_update")
def webhook_fundamental_update(update: FundamentalUpdate, authorization: Optional[str] = Header(None)):
    """
    Ruta que recibe la Miel (booleanos extraídos por n8n) y actualiza la matriz.
    """
    verificar_token(authorization)
    
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        raise HTTPException(status_code=503, detail="Firebase no inicializado")
        
    try:
        activo_normalizado = normalizar_activo(update.activo)
        doc_ref = db.collection("trading_matrix").document(activo_normalizado)
        doc = doc_ref.get()
        
        if not doc.exists:
            raise HTTPException(status_code=404, detail=f"El activo {activo_normalizado} no existe")
            
        data = doc.to_dict()
        
        if "confirmaciones_fundamentales" not in data:
            data["confirmaciones_fundamentales"] = {}
            
        if update.noticias_impacto_favorables is not None:
            data["confirmaciones_fundamentales"]["noticias_impacto_favorables"] = update.noticias_impacto_favorables
            
        if update.ipo_liquidez_positiva is not None:
            data["confirmaciones_fundamentales"]["ipo_liquidez_positiva"] = update.ipo_liquidez_positiva
            
        if update.spo_liquidez_positiva is not None:
            data["confirmaciones_fundamentales"]["spo_liquidez_positiva"] = update.spo_liquidez_positiva
        
        # Recalcular score
        score = recalcular_score_ponderado(data)
        data["score_porcentaje"] = round(score, 2)
        data["gatillo_entrada"] = score >= 80.0
        
        if data["gatillo_entrada"] and data.get("estado_ejecucion", "INACTIVO") == "INACTIVO":
            data["estado_ejecucion"] = "PENDIENTE_EJECUCIÓN"
            print(f"| SEMÁFORO | {activo_normalizado} ha cambiado a PENDIENTE_EJECUCIÓN (Vía Fundamental)")
            notificar_botpress_mia(activo_normalizado, data)
            
        data["ultimo_update"] = datetime.datetime.now(datetime.timezone.utc).isoformat() if hasattr(datetime, "timezone") else datetime.datetime.now().isoformat()
        
        doc_ref.set(data)
        print(f"| FIREBASE SUCCESS | Confirmaciones fundamentales actualizadas. Score: {data['score_porcentaje']}%")
        
        return {
            "status": "success",
            "activo": activo_normalizado,
            "estado_ejecucion": data.get("estado_ejecucion")
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"| CLOUD ERROR | Error en webhook_fundamental_update: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/test_rss_llm_polling")
def test_rss_llm_polling(authorization: Optional[str] = Header(None)):
    """
    Simulates polling of the XML RSS feed and queries configured LLMs
    (Gemini, ChatGPT, Grok) with the XML content to test their parsing/analysis capacity.
    """
    verificar_token(authorization)
    
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        raise HTTPException(status_code=503, detail="Firebase no inicializado")
        
    try:
        # 1. Fetch XML feed content
        docs = db.collection("trading_matrix").stream()
        xml_items = []
        for doc in docs:
            activo_id = doc.id
            data = doc.to_dict()
            apoyo = data.get("aprendizaje_mia", {})
            
            raw_sent = apoyo.get('sentimiento_acumulado', 'NEUTRAL').upper()
            if raw_sent == "BULLISH":
                sentimiento_val = "True"
            elif raw_sent == "BEARISH":
                sentimiento_val = "False"
            else:
                sentimiento_val = "NEUTRAL"
                
            item_xml = f"""
        <activo name="{activo_id}">
            <trades_totales>{apoyo.get('trades_totales', 0)}</trades_totales>
            <trades_ganados>{apoyo.get('trades_ganados', 0)}</trades_ganados>
            <win_rate_historico>{apoyo.get('win_rate_historico', 50.0)}</win_rate_historico>
            <racha_actual>{apoyo.get('racha_actual', 0)}</racha_actual>
            <sentimiento_acumulado>{sentimiento_val}</sentimiento_acumulado>
            <factor_ajuste_probabilidad>{apoyo.get('factor_ajuste_probabilidad', 0.0)}</factor_ajuste_probabilidad>
            <ultimo_update>{data.get('ultimo_update', '')}</ultimo_update>
            <score_porcentaje>{data.get('score_porcentaje', 0.0)}</score_porcentaje>
        </activo>"""
            xml_items.append(item_xml)
            
        xml_content = f"""<?xml version="1.0" encoding="UTF-8"?>
<mia_trading_learnings>
    <canal>
        <titulo>MIA Trading Bot learnings</titulo>
        <descripcion>Base de conocimiento (KB) de Mia en Trading Algoritmico</descripcion>
        <generacion>{datetime.datetime.now(datetime.timezone.utc).isoformat() if hasattr(datetime, 'timezone') else datetime.datetime.now().isoformat()}</generacion>
        <activos>{"".join(xml_items)}
        </activos>
    </canal>
</mia_trading_learnings>"""

        # 2. Run LLM tests
        responses = {}
        active_llms = []
        
        # Test Gemini
        if GEMINI_API_KEY and GEMINI_API_KEY != "TU_LLAVE_DE_GEMINI":
            active_llms.append("Gemini")
            responses["Gemini"] = consultar_llm_rss_helper(xml_content, "gemini")
            
        # Test ChatGPT (OpenAI)
        if OPENAI_API_KEY and OPENAI_API_KEY != "TU_LLAVE_DE_OPENAI":
            active_llms.append("ChatGPT")
            responses["ChatGPT"] = consultar_llm_rss_helper(xml_content, "chatgpt")
            
        # Test Grok
        if GROK_API_KEY and GROK_API_KEY != "TU_LLAVE_DE_GROK":
            active_llms.append("Grok")
            responses["Grok"] = consultar_llm_rss_helper(xml_content, "grok")
            
        return {
            "status": "success",
            "active_llms": active_llms,
            "xml_preview": xml_content[:400] + "...",
            "llm_responses": responses
        }
    except Exception as e:
        print(f"| TEST RSS LLM ERROR | {e}")
        raise HTTPException(status_code=500, detail=str(e))

class MetaApiExecution(BaseModel):
    activo: str
    ticket: str
    accion: str = ""
    score: float
    precio_ejecucion: float
    stop_loss: Optional[float] = 0.0
    take_profit: Optional[float] = 0.0
    ejecutada_mt5: bool = True
    motivo: str = "Cumple parámetros de matriz técnica y de riesgo"

@app.post("/webhook_marcar_ejecutado")
def webhook_marcar_ejecutado(ejecucion: MetaApiExecution, authorization: Optional[str] = Header(None)):
    """
    Recibe la confirmación desde Botpress (MetaApi) de que el trade se ha ejecutado.
    Cambia el estado a EJECUTADO, llama a la KB, y genera el log de auditoría inmutable.
    """
    verificar_token(authorization)
    invalidar_cache_dashboard()
    
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        raise HTTPException(status_code=503, detail="Firebase no inicializado")
        
    activo_norm = normalizar_activo(ejecucion.activo)
    doc_ref = db.collection("trading_matrix").document(activo_norm)
    
    try:
        data = doc_ref.get().to_dict() or {}
        data["estado_ejecucion"] = "EJECUTADO"
        doc_ref.set(data, merge=True)
        
        # Generar Log .txt
        import os
        from datetime import datetime
        log_dir = "logs"
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)
            
        log_path = os.path.join(log_dir, "trading_audit_log.txt")
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[{fecha}] TICKET: {ejecucion.ticket} | ACTIVO: {ejecucion.activo} | SCORE: {ejecucion.score}% | PRECIO: {ejecucion.precio_ejecucion}\n")
            
        # Enriquecer log con detalles de confirmaciones de la matriz
        activas = [k.replace("_", " ").upper() for k, v in data.get("confirmaciones_tecnicas", {}).items() if isinstance(v, bool) and v]
        confirmaciones_str = " + ".join(activas) if activas else "Setup Base"
        
        utc_hour = datetime.now(timezone.utc).hour if hasattr(datetime, "timezone") else datetime.now().hour
        sesion = "NY"
        if 0 <= utc_hour < 7: sesion = "ASIA"
        elif 7 <= utc_hour < 12: sesion = "LONDRES"
        
        estrategia_base = "SMC Setup"
        str_ejecutada = "SÍ" if ejecucion.ejecutada_mt5 else "NO"
        
        detalle_str = f"{ejecucion.activo} | {fecha} | {sesion} | {estrategia_base} | {confirmaciones_str} | SCORE: {ejecucion.score}% | EJECUTADA EN MT5: {str_ejecutada} | MOTIVO: {ejecucion.motivo}"
        
        audit_ref = db.collection("mia_audit_logs").document(str(ejecucion.ticket))
        audit_ref.set({
            "ticket": ejecucion.ticket,
            "activo": ejecucion.activo,
            "accion": ejecucion.accion,
            "score": ejecucion.score,
            "precio_ejecucion": ejecucion.precio_ejecucion,
            "stop_loss": ejecucion.stop_loss,
            "take_profit": ejecucion.take_profit,
            "tp": ejecucion.take_profit,
            "sl": ejecucion.stop_loss,
            "fecha": fecha,
            "timestamp": datetime.now().isoformat(),
            "detalle_setup": detalle_str,
            "confirmaciones_tecnicas": data.get("confirmaciones_tecnicas", {}),
            "confirmaciones_fundamentales": data.get("confirmaciones_fundamentales", {}),
            "confirmaciones_institucionales": data.get("confirmaciones_institucionales", {})
        }, merge=True)
            
        print(f"| AUDITORÍA | Trade registrado en TXT y Firebase (mia_audit_logs) para {ejecucion.activo}")
        
        return {"status": "success", "mensaje": "Trade ejecutado y auditado en TXT y Firebase"}
    except Exception as e:
        print(f"| AUDITORÍA ERROR | {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/webhook_marcar_rechazado")
def webhook_marcar_rechazado(payload: dict, authorization: Optional[str] = Header(None)):
    """
    Actualiza el Live Feed (mia_audit_logs) indicando el motivo exacto por el cual 
    el cerebro o el MetaAPI rechazó la orden.
    """
    verificar_token(authorization)
    invalidar_cache_dashboard()
    
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        raise HTTPException(status_code=503, detail="Firebase no inicializado")
        
    activo = payload.get("activo", "")
    motivo = payload.get("motivo", "Rechazado")
    activo_norm = normalizar_activo(activo)
    
    try:
        # 1. Resetear el semáforo en trading_matrix para que pueda volver a intentarlo en el futuro
        doc_matrix_ref = db.collection("trading_matrix").document(activo_norm)
        matrix_data = doc_matrix_ref.get().to_dict() or {}
        if matrix_data.get("estado_ejecucion") == "EJECUTADO":
            matrix_data["estado_ejecucion"] = "INACTIVO"
            doc_matrix_ref.set(matrix_data, merge=True)
            print(f"| SEMÁFORO | Reset a INACTIVO para {activo_norm} debido a rechazo de MetaAPI/Killzone.")
            
        # 2. Buscar el registro EVAL más reciente de este activo y actualizar su motivo
        docs = db.collection("mia_audit_logs").order_by("timestamp", direction=firestore.Query.DESCENDING).limit(20).stream()
        for doc in docs:
            data = doc.to_dict()
            if normalizar_activo(data.get("activo", "")) == activo_norm and not data.get("ejecutada_mt5", False):
                data["motivo"] = motivo
                detalle = data.get("detalle_setup", "")
                if "MOTIVO: " in detalle:
                    detalle = detalle.split("MOTIVO: ")[0] + f"MOTIVO: {motivo}"
                    data["detalle_setup"] = detalle
                    
                db.collection("mia_audit_logs").document(doc.id).set(data, merge=True)
                print(f"| AUDITORÍA | Motivo de rechazo actualizado para {activo_norm}: {motivo}")
                
                # Actualizar Caché Global en RAM
                global GLOBAL_AUDIT_LOGS
                if GLOBAL_AUDIT_LOGS is not None:
                    for i, log in enumerate(GLOBAL_AUDIT_LOGS):
                        if log.get("ticket") == str(data.get("ticket")) or (log.get("activo") == data.get("activo") and not log.get("ejecutada_mt5")):
                            GLOBAL_AUDIT_LOGS[i] = data
                            break
                            
                break
                
        return {"status": "success", "mensaje": "Motivo de rechazo actualizado"}
    except Exception as e:
        print(f"| AUDITORÍA ERROR | Error al actualizar rechazo: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/webhook_marcar_parcial")
def webhook_marcar_parcial(ejecucion: MetaApiExecution, authorization: Optional[str] = Header(None)):
    """
    Recibe la confirmación desde Botpress (MetaApi) de que el CIERRE PARCIAL se ha ejecutado.
    Actualiza la lógica booleana en Firebase.
    """
    verificar_token(authorization)
    
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        raise HTTPException(status_code=503, detail="Firebase no inicializado")
        
    activo_norm = normalizar_activo(ejecucion.activo)
    doc_ref = db.collection("trading_matrix").document(activo_norm)
    
    try:
        data = doc_ref.get().to_dict() or {}
        data["estado_ejecucion"] = "PARCIAL_CERRADO"
        data["parcial_tomado"] = True
        doc_ref.set(data, merge=True)
        
        # Generar Log en Firebase
        fecha = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        audit_data = {
            "tipo": "CIERRE_PARCIAL_80",
            "ticket": ejecucion.ticket,
            "activo": ejecucion.activo,
            "score_confluencias": ejecucion.score,
            "precio_ejecucion": ejecucion.precio_ejecucion,
            "fecha": fecha,
            "timestamp": datetime.datetime.now().isoformat()
        }
        audit_ref = db.collection("mia_audit_logs").document(f"PARCIAL_{ejecucion.ticket}_{ejecucion.activo}")
        audit_ref.set(audit_data)
            
        print(f"| AUDITORÍA PARCIAL | Cierre Parcial registrado en Firebase para {ejecucion.activo}")
        
        # Actualizar Caché Global en RAM
        global GLOBAL_AUDIT_LOGS
        if GLOBAL_AUDIT_LOGS is not None:
            GLOBAL_AUDIT_LOGS.append(audit_data)
            invalidar_cache_dashboard()
        
        return {"status": "success", "mensaje": "Cierre Parcial auditado en Firebase"}
    except Exception as e:
        print(f"| AUDITORÍA ERROR | {e}")
        raise HTTPException(status_code=500, detail=str(e))

class UpdateBalancePayload(BaseModel):
    balance: float
    equity: float = 0.0
    floating_pnl: float = 0.0

@app.post("/webhook_update_balance")
def webhook_update_balance(payload: UpdateBalancePayload, authorization: Optional[str] = Header(None)):
    """
    Recibe el balance en vivo desde el MT5 Executor (Nube) y lo guarda en Firebase para el Dashboard.
    """
    verificar_token(authorization)
    
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        raise HTTPException(status_code=503, detail="Firebase no inicializado")
        
    try:
        from datetime import datetime
        db.collection("system_memory").document("broker_state").set({
            "live_balance": payload.balance,
            "equity": payload.equity,
            "floating_pnl": payload.floating_pnl,
            "timestamp": datetime.now().isoformat()
        }, merge=True)
        
        invalidar_cache_dashboard()
        return {"status": "success", "mensaje": "Balance actualizado"}
    except Exception as e:
        print(f"| GESTOR BALANCE ERROR | {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/pnl_hoy")
def api_pnl_hoy(authorization: Optional[str] = Header(None)):
    """
    Devuelve la suma total del PNL de todas las operaciones cerradas el día de hoy.
    """
    verificar_token(authorization)
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        raise HTTPException(status_code=503, detail="Firebase no inicializado")
        
    try:
        from datetime import datetime
        hoy_str = datetime.now().strftime("%Y-%m-%d")
        
        asegurar_cache_firebase()
        global GLOBAL_AUDIT_LOGS
        
        pnl_total = 0.0
        if GLOBAL_AUDIT_LOGS:
            for data in GLOBAL_AUDIT_LOGS:
                fecha_doc = data.get("fecha", "")
                if fecha_doc.startswith(hoy_str):
                    # Solo sumar si es un CIERRE_TOTAL (o PARCIAL si se incluye)
                    accion = data.get("accion", "")
                    if accion in ["CIERRE_TOTAL", "CIERRE_PARCIAL_80", "CIERRE_PARCIAL"]:
                        pnl_total += float(data.get("pnl", 0.0))
                    
        return {"status": "success", "pnl_hoy": pnl_total}
    except Exception as e:
        print(f"| API ERROR | Error calculando PNL de hoy: {e}")
        return {"status": "error", "pnl_hoy": 0.0, "detalle": str(e)}


@app.get("/resumen_trades_hoy")
def resumen_trades_hoy(authorization: Optional[str] = Header(None)):
    """
    Consulta la base de datos de auditoría de Firebase (mia_audit_logs)
    y devuelve un resumen formateado de los trades ejecutados el día de hoy
    para que Botpress pueda mostrarlo en el chat.
    """
    verificar_token(authorization)
    
    global firebase_inicializado, db
    if not firebase_inicializado or db is None:
        raise HTTPException(status_code=503, detail="Firebase no inicializado")
        
    try:
        from datetime import datetime, timezone
        import pytz
        
        # Obtener la fecha de hoy en formato YYYY-MM-DD
        hoy_str = datetime.now().strftime("%Y-%m-%d")
        
        asegurar_cache_firebase()
        global GLOBAL_AUDIT_LOGS
        
        trades_hoy = []
        if GLOBAL_AUDIT_LOGS:
            # Ordenamos por timestamp descendente simulando la base de datos
            logs_ordenados = sorted(GLOBAL_AUDIT_LOGS, key=lambda x: x.get("timestamp", ""), reverse=True)
            for data in logs_ordenados:
                fecha_doc = data.get("fecha", "")
                if fecha_doc.startswith(hoy_str):
                    trades_hoy.append(data)
                
        if len(trades_hoy) == 0:
            return {"status": "success", "mensaje_chat": f"Padre, hoy ({hoy_str}) no hemos ejecutado ningún trade todavía. Sigo escaneando el mercado pacientemente."}
            
        resumen = f"Padre, este es el resumen de hoy ({hoy_str}):\n\n"
        for t in trades_hoy:
            tipo = t.get("tipo", "EJECUCIÓN")
            activo = t.get("activo", "DESCONOCIDO")
            score = t.get("score_confluencias", t.get("score", 0))
            resumen += f"• [{tipo}] {activo} | Score: {score}%\n"
            
        resumen += f"\nTotal de movimientos hoy: {len(trades_hoy)}."
        
        return {"status": "success", "mensaje_chat": resumen}
        
    except Exception as e:
        print(f"| RESUMEN ERROR | Error al generar resumen de trades: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ------------------------------------------------------------------------------
# DASHBOARD INSTITUCIONAL
# ------------------------------------------------------------------------------
@app.get("/dashboard", response_class=HTMLResponse)
async def render_dashboard():
    try:
        with open("dashboard_mia.html", "r", encoding="utf-8") as f:
            html_content = f.read()
        return HTMLResponse(content=html_content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"No se pudo cargar el dashboard: {e}")
GLOBAL_AUDIT_LOGS = None
GLOBAL_SYSTEM_LOGS = None
GLOBAL_PATRONES = None
GLOBAL_MATRICES = None
GLOBAL_MIA_COLLECTIVE = None
GLOBAL_INDICADORES = None
ULTIMO_FETCH_FIREBASE = None
# Caché RAM de activos ya matriculados en Firebase. Evita lecturas repetidas a Firestore.
# Se llena en startup y se actualiza cuando se detecta un activo nuevo.
ACTIVOS_INICIALIZADOS: set = set()

def asegurar_cache_firebase():
    global firebase_inicializado, db
    global GLOBAL_AUDIT_LOGS, GLOBAL_SYSTEM_LOGS, GLOBAL_PATRONES, GLOBAL_MATRICES, ULTIMO_FETCH_FIREBASE
    global GLOBAL_MIA_COLLECTIVE, GLOBAL_INDICADORES
    
    if not firebase_inicializado or db is None:
        return
        
    from datetime import datetime
    ahora = datetime.now()
    
    # Refrescar caché de Firebase cada 5 minutos para evitar agotar la cuota de lectura (Quota Exceeded / Resource Exhausted)
    necesita_refresh = False
    if ULTIMO_FETCH_FIREBASE is None or GLOBAL_AUDIT_LOGS is None:
        necesita_refresh = True
    elif (ahora - ULTIMO_FETCH_FIREBASE).total_seconds() > 300.0:
        necesita_refresh = True
        
    if necesita_refresh:
        try:
            # system_logs
            sys_logs = db.collection("mia_system_logs").order_by("timestamp", direction=firestore.Query.DESCENDING).limit(10).stream()
            GLOBAL_SYSTEM_LOGS = [sl.to_dict() for sl in sys_logs]
            
            # mia_audit_logs (Limitamos a 1500 para evitar desbordar la memoria RAM de Railway o causar Timeout)
            logs = db.collection("mia_audit_logs").order_by("timestamp", direction=firestore.Query.DESCENDING).limit(1500).stream()
            GLOBAL_AUDIT_LOGS = [l.to_dict() for l in logs]
            
            # trading_matrix
            matrices = db.collection("trading_matrix").stream()
            GLOBAL_MATRICES = {m.id: m.to_dict().get("score_porcentaje", 0) for m in matrices}
            
            # mia_kb / patrones
            patrones = db.collection("mia_kb").document("patrones_ict_smc").collection("detalle").stream()
            GLOBAL_PATRONES = [p.to_dict() for p in patrones]
            
            # mia_collective
            try:
                mem_doc = db.collection("system_memory").document("mia_collective").get()
                GLOBAL_MIA_COLLECTIVE = mem_doc.to_dict() if mem_doc.exists else {}
            except:
                GLOBAL_MIA_COLLECTIVE = {}
                
            # indicadores_impacto
            indicadores = db.collection("mia_kb").document("indicadores_impacto").collection("detalle").stream()
            GLOBAL_INDICADORES = [{"nombre": ind.id, **ind.to_dict()} for ind in indicadores]
            
            ULTIMO_FETCH_FIREBASE = ahora
        except Exception as fe:
            print(f"| FIREBASE CACHE ERROR | Error recargando caché: {fe}")
            if GLOBAL_AUDIT_LOGS is None: GLOBAL_AUDIT_LOGS = []
            if GLOBAL_SYSTEM_LOGS is None: GLOBAL_SYSTEM_LOGS = []
            if GLOBAL_PATRONES is None: GLOBAL_PATRONES = []
            if GLOBAL_MATRICES is None: GLOBAL_MATRICES = {}

@app.get('/api/export_trades')
def api_export_trades():
    global GLOBAL_AUDIT_LOGS
    try:
        asegurar_cache_firebase()
    except: pass
    return {"status": "success", "data": GLOBAL_AUDIT_LOGS}

@app.get("/api/dashboard_data")
def api_dashboard_data():
    """Devuelve los datos estructurados para renderizar el Dashboard."""
    global firebase_inicializado, db
    global GLOBAL_AUDIT_LOGS, GLOBAL_SYSTEM_LOGS, GLOBAL_PATRONES, GLOBAL_MATRICES, ULTIMO_FETCH_FIREBASE
    global DASHBOARD_CACHE_DATA, DASHBOARD_CACHE_TIME
    
    if not firebase_inicializado or db is None:
        return {"status": "error", "message": "Firebase no inicializado"}

    # Caché en RAM de 3 minutos para el bloque completo del dashboard para proteger la cuota de Firebase
    ahora_t = time.time()
    if DASHBOARD_CACHE_DATA and (ahora_t - DASHBOARD_CACHE_TIME) < 180.0:
        # Devolver datos de caché RAM directamente sin lecturas
        DASHBOARD_CACHE_DATA["recent_logs"] = GLOBAL_AUDIT_LOGS
        return {"status": "success", "data": DASHBOARD_CACHE_DATA, "cached": True}

    data = {
        "balance_base": 5000.0,
        "balance_actual": 5000.0,
        "equity": 5000.0,
        "floating_pnl": 0.0,
        "pnl_total": 0.0,
        "kpis": {},
        "rendimiento_activos": {},
        "curva_equity": [],
        "estrategias": [],
        "killzones": [],
        "indicadores": [],
        "matriz_scores": {},
        "feed": [],
        "operaciones_activas": [],
        "system_logs": []
    }

    try:
        asegurar_cache_firebase()

        data["system_logs"] = GLOBAL_SYSTEM_LOGS
        
        balance_actual = None
        equity_actual = None
        floating_pnl = 0.0
        try:
            broker_doc = db.collection("system_memory").document("broker_state").get()
            if broker_doc.exists:
                broker_data = broker_doc.to_dict()
                balance_actual = float(broker_data.get("live_balance", 5000.0))
                equity_actual = float(broker_data.get("equity", balance_actual))
                floating_pnl = float(broker_data.get("floating_pnl", 0.0))
        except: pass
        
        data["floating_pnl"] = floating_pnl
        
        from datetime import datetime
        hoy_str = datetime.now().strftime("%Y-%m-%d")
        
        activos_stats = {}
        todos_los_logs = []
        todas_las_entradas = []
        
        tickets_cerrados = set()
        dict_activas = {}
        parciales_tomados = 0
        
        for l in GLOBAL_AUDIT_LOGS:
            accion = l.get("accion", "")
            ticket = l.get("ticket")
            
            if accion in ["CIERRE_TOTAL", "CIERRE_PARCIAL"]:
                todos_los_logs.append(l)
                if accion == "CIERRE_TOTAL" and ticket:
                    tickets_cerrados.add(ticket)
                if accion == "CIERRE_PARCIAL":
                    parciales_tomados += 1
                    todas_las_entradas.append(l)
            else:
                todas_las_entradas.append(l)
                if ticket and accion in ["COMPRA", "VENTA"]:
                    # Guardamos la mas reciente si hubiera duplicados
                    if ticket not in dict_activas or l.get("timestamp", "") > dict_activas[ticket].get("timestamp", ""):
                        dict_activas[ticket] = l
                        
        operaciones_activas = []
        for t, d in dict_activas.items():
            if t not in tickets_cerrados:
                operaciones_activas.append(d)
                
        data["operaciones_activas"] = operaciones_activas
                
        # Procesar Feed de Oportunidades (Últimas 500)
        todas_las_entradas = sorted(todas_las_entradas, key=lambda x: x.get("timestamp", ""), reverse=True)[:500]
        for e in todas_las_entradas:
            score_val = float(e.get("score", e.get("score_confluencias", 0)))
            detalle = e.get("detalle_setup")
            if not detalle:
                activo = e.get('activo', 'UNKNOWN')
                if activo == "SYSTEM": 
                    activo = "EURUSD"
                
                fecha = e.get('fecha', '')
                sesion = "NUEVA_YORK"
                if fecha:
                    try:
                        from datetime import datetime
                        dt = datetime.strptime(fecha, "%Y-%m-%d %H:%M:%S")
                        h = dt.hour
                        if 1 <= h < 6: sesion = "LONDRES"
                        elif 6 <= h < 16: sesion = "NUEVA_YORK"
                        else: sesion = "ASIA"
                    except:
                        pass
                
                estrategia = e.get('estrategia', 'SMC Setup')
                if not estrategia or estrategia.strip() == "":
                    estrategia = "SMC Setup"
                    
                detalle = f"{activo} | {fecha} | {sesion} | {estrategia} | EVALUACIÓN | SCORE: {score_val}% | EJECUTADA EN MT5: {('SÍ' if e.get('ejecutada_mt5') else 'NO')} | MOTIVO: {e.get('motivo', '')}"
            
            data["feed"].append({
                "texto": detalle,
                "color": "#00e68a" if score_val >= 80 else "#00b4d8",
                "timestamp": e.get("timestamp", "") or e.get("fecha", "")
            })
                
        # Ordenar por timestamp
        todos_los_logs = sorted(todos_los_logs, key=lambda x: x.get("timestamp", ""))

        base_assets = ["XAUUSD", "EURUSD", "GBPJPY", "AUDUSD", "GBPUSD", "NZDCAD"]
        activos_stats = {
            a: {"hoy": 0, "semana": 0, "mes": 0, "trimestre": 0, "semestre": 0, "anual": 0, "trades": 0, "pnl_total": 0}
            for a in base_assets
        }
        
        from datetime import datetime, timedelta
        now = datetime.now()
        semana_atras = now - timedelta(days=7)
        mes_atras = now - timedelta(days=30)
        trim_atras = now - timedelta(days=90)
        sem_atras = now - timedelta(days=180)
        anio_atras = now - timedelta(days=365)
        
        # Calcular el balance de la curva de forma retroactiva para que el ultimo punto coincida con el balance actual real
        pnl_sum_total = sum(l.get("pnl", 0.0) for l in todos_los_logs)
        if balance_actual is not None:
            balance_inicial_curva = balance_actual - pnl_sum_total
        else:
            balance_inicial_curva = 5000.0
            balance_actual = balance_inicial_curva + pnl_sum_total
            equity_actual = balance_actual + floating_pnl
            
        data["balance_actual"] = balance_actual
        data["equity"] = equity_actual
        
        balance_curva = balance_inicial_curva
        data["balance_base"] = balance_inicial_curva
        
        for l in todos_los_logs:
            pnl = l.get("pnl", 0.0)
            activo = l.get("activo", "UNKNOWN")
            fecha_str = l.get("fecha", "")
            
            data["pnl_total"] += pnl
            balance_curva += pnl
            
            data["curva_equity"].append({
                "fecha": fecha_str,
                "balance": balance_curva
            })

            if activo not in activos_stats:
                activos_stats[activo] = {"hoy": 0, "semana": 0, "mes": 0, "trimestre": 0, "semestre": 0, "anual": 0, "trades": 0, "pnl_total": 0}
            
            activos_stats[activo]["trades"] += 1
            activos_stats[activo]["pnl_total"] += pnl
            if fecha_str.startswith(hoy_str):
                activos_stats[activo]["hoy"] += pnl
                
            if fecha_str:
                try:
                    dt = datetime.strptime(fecha_str, "%Y-%m-%d %H:%M:%S")
                    if dt >= anio_atras: activos_stats[activo]["anual"] += pnl
                    if dt >= sem_atras: activos_stats[activo]["semestre"] += pnl
                    if dt >= trim_atras: activos_stats[activo]["trimestre"] += pnl
                    if dt >= mes_atras: activos_stats[activo]["mes"] += pnl
                    if dt >= semana_atras: activos_stats[activo]["semana"] += pnl
                except: pass

        data["rendimiento_activos"] = activos_stats

        # Calcular KPIs dinámicos con clasificación de cierres por ticket
        from collections import defaultdict
        logs_por_ticket = defaultdict(list)
        if GLOBAL_AUDIT_LOGS:
            for l in GLOBAL_AUDIT_LOGS:
                t = str(l.get("ticket", ""))
                if t and t != "0" and t != "None":
                    logs_por_ticket[t].append(l)

        total_tp = 0
        total_sl = 0
        total_be = 0
        total_manual_parcial = 0
        total_manual_directo = 0
        
        for t, t_logs in logs_por_ticket.items():
            acciones = [str(l.get("accion", "")).upper() for l in t_logs]
            pnls = [float(l.get("pnl", 0.0)) for l in t_logs]
            comentarios = [str(l.get("estrategia", "")).upper() for l in t_logs]
            
            has_cierre_total = "CIERRE_TOTAL" in acciones
            has_cierre_parcial = "CIERRE_PARCIAL" in acciones or any("PARCIAL" in c for c in comentarios)
            
            if not has_cierre_total:
                continue
                
            if has_cierre_parcial:
                cierre_total_log = next((l for l in t_logs if str(l.get("accion")).upper() == "CIERRE_TOTAL"), None)
                if cierre_total_log:
                    pnl_cierre = float(cierre_total_log.get("pnl", 0.0))
                    # Si el PnL del cierre final es cercano a 0 (BE), lo contamos como BE
                    if abs(pnl_cierre) <= 1.5:
                        total_be += 1
                    else:
                        total_manual_parcial += 1
                else:
                    total_be += 1
            else:
                cierre_total_log = next((l for l in t_logs if str(l.get("accion")).upper() == "CIERRE_TOTAL"), None)
                if cierre_total_log:
                    ct_comentario = str(cierre_total_log.get("estrategia", "")).upper()
                    pnl_cierre = float(cierre_total_log.get("pnl", 0.0))
                    
                    if pnl_cierre < 0:
                        total_sl += 1
                    elif "DESAPARICION" in ct_comentario or "MANUAL" in ct_comentario:
                        total_manual_directo += 1
                    else:
                        total_tp += 1
                else:
                    pnl_final = sum(pnls)
                    if pnl_final < 0:
                        total_sl += 1
                    else:
                        total_tp += 1

        ganados = len([l for l in todos_los_logs if l.get("pnl", 0) > 0])
        total_cerrados = len(todos_los_logs)
        win_rate = round((ganados / total_cerrados * 100), 2) if total_cerrados > 0 else 0
        
        # Identificar los verdaderos trades ejecutados (no EVALs)
        verdaderos_trades = [t for t in todas_las_entradas if t.get("ejecutada_mt5") == True or t.get("accion") in ["COMPRA", "VENTA", "CIERRE_PARCIAL", "CIERRE_TOTAL"] or "Ejecutada por Escáner Cloud" in str(t.get("detalle_setup", ""))]
        total_trades = len(verdaderos_trades) + total_cerrados
        
        # Integrar mia_collective con KPIs calculados
        m = GLOBAL_MIA_COLLECTIVE if GLOBAL_MIA_COLLECTIVE else {}
        data["kpis"] = {
            "win_rate": win_rate,
            "total_trades": total_trades,
            "patron_estrella": m.get("patron_estrella_ict_smc", "-"),
            "patron_estrella_wr": m.get("patron_estrella_win_rate", 0),
            "parciales_tomados": parciales_tomados,
            "total_tp": total_tp,
            "total_sl": total_sl,
            "total_be": total_be,
            "total_manual_parcial": total_manual_parcial,
            "total_manual_directo": total_manual_directo
        }
        
        # 2. trading_matrix (Scores en vivo)
        for m_id, score_p in GLOBAL_MATRICES.items():
            data["matriz_scores"][m_id] = score_p

        # 4. Estrategias (mia_kb/patrones_ict_smc)
        for pdata in GLOBAL_PATRONES:
            if pdata.get("ocurrencias", 0) > 0:
                # Tratamos de recuperar el ID original (nombre) pero no lo tenemos en el dict a menos que lo guardemos.
                # Como un hack, usaremos patron_estrella si coincide
                data["estrategias"].append({
                    "nombre": pdata.get("nombre", "Patrón").replace("_", " + "),
                    "win_rate": pdata.get("win_rate", 0),
                    "ocurrencias": pdata.get("ocurrencias", 0)
                })

        data["estrategias"] = sorted(data["estrategias"], key=lambda x: x["win_rate"], reverse=True)

        # 5. Killzones Dinámicas (Calculadas a partir de todos_los_logs)
        killzone_stats = {
            "LONDRES": {"ganados": 0, "perdidos": 0},
            "NUEVA_YORK": {"ganados": 0, "perdidos": 0},
            "ASIA": {"ganados": 0, "perdidos": 0}
        }
        
        for l in todos_los_logs:
            pnl = l.get("pnl", 0.0)
            fecha = l.get("fecha", "")
            ses = "NUEVA_YORK"
            if fecha:
                try:
                    dt = datetime.strptime(fecha, "%Y-%m-%d %H:%M:%S")
                    h = dt.hour
                    if 1 <= h < 6: ses = "LONDRES"
                    elif 6 <= h < 16: ses = "NUEVA_YORK"
                    else: ses = "ASIA"
                except: pass
                
            if pnl > 0:
                killzone_stats[ses]["ganados"] += 1
            else:
                killzone_stats[ses]["perdidos"] += 1
                
        data["killzones"] = []
        for kz, stats in killzone_stats.items():
            tot = stats["ganados"] + stats["perdidos"]
            if tot > 0:
                data["killzones"].append({
                    "nombre": kz,
                    "win_rate": round(stats["ganados"] / tot * 100, 2),
                    "trades": tot,
                    "ganados": stats["ganados"],
                    "perdidos": stats["perdidos"]
                })
        data["killzones"] = sorted(data["killzones"], key=lambda x: x["win_rate"], reverse=True)
        data["rendimiento_sesiones"] = data["killzones"]

        # 6. Indicadores/Ponderaciones (mia_kb/indicadores_impacto)
        if GLOBAL_INDICADORES:
            for idata in GLOBAL_INDICADORES:
                if idata.get("trades_con_indicador", 0) > 0:
                    data["indicadores"].append({
                        "nombre": idata.get("nombre"),
                        "win_rate": idata.get("win_rate_indicador", 0),
                        "trades": idata.get("trades_con_indicador", 0)
                    })

        data["indicadores"] = sorted(data["indicadores"], key=lambda x: x["win_rate"], reverse=True)

        # Actualizar la caché global
        DASHBOARD_CACHE_DATA = data
        DASHBOARD_CACHE_TIME = time.time()
        print("| CACHE | Datos del dashboard actualizados y guardados en memoria")

        data["recent_logs"] = GLOBAL_AUDIT_LOGS
        return {"status": "success", "data": data}

    except Exception as e:
        print(f"| API ERROR | Fallo al recopilar datos del dashboard: {e}")
        # Si falla por 429 u otro error, intentar devolver la caché antigua si existe
        if DASHBOARD_CACHE_DATA:
            print("| CACHE FALLBACK | Sirviendo datos antiguos por fallo en Firebase")
            return {"status": "success", "data": DASHBOARD_CACHE_DATA, "warning": str(e)}
        return {"status": "error", "message": str(e)}

@app.get("/api/open_trades")
def api_open_trades():
    """
    Retorna la lista de tickets que están activos en Firebase (COMPRA/VENTA)
    y que aún no han sido cerrados. Lee directo de la memoria RAM.
    """
    if GLOBAL_AUDIT_LOGS is None:
        asegurar_cache_firebase()
    
    open_tickets = []
    if GLOBAL_AUDIT_LOGS:
        for l in GLOBAL_AUDIT_LOGS:
            if l.get("accion") in ["COMPRA", "VENTA"] and l.get("ticket"):
                open_tickets.append(str(l.get("ticket")))
                
    return {"status": "success", "open_tickets": open_tickets}

@app.get("/api/get_trade_tp/{ticket}")
def get_trade_tp(ticket: str):
    """
    Busca en mia_audit_logs el log original del ticket para devolver su take_profit original
    y si ya tiene registrado un cierre parcial en Firebase.
    """
    if not firebase_inicializado or db is None:
        return {"status": "error", "message": "Firebase no inicializado"}
    try:
        doc = db.collection("mia_audit_logs").document(str(ticket)).get()
        tp = 0.0
        if doc.exists:
            data = doc.to_dict()
            tp = data.get("take_profit", data.get("tp", 0.0))
            
        if not tp:
            # Buscar en trading_alerts si no está en mia_audit_logs
            alerts = db.collection("trading_alerts").where("ticket", "==", int(ticket)).limit(1).stream()
            for a in alerts:
                tp = a.to_dict().get("take_profit", 0.0)
                
        # Buscar si ya se tomó un parcial (si existe un documento que empiece por PARCIAL_{ticket} o un log con accion CIERRE_PARCIAL)
        parcial_tomado = False
        if GLOBAL_AUDIT_LOGS:
            for l in GLOBAL_AUDIT_LOGS:
                if str(l.get("ticket")) == str(ticket) and l.get("accion") == "CIERRE_PARCIAL":
                    parcial_tomado = True
                    break
        
        # Fallback de búsqueda directa en Firestore si no está en caché
        if not parcial_tomado:
            p_doc = db.collection("mia_audit_logs").document(f"PARCIAL_{ticket}").get()
            if p_doc.exists:
                parcial_tomado = True
                
        return {"status": "success", "tp": float(tp), "parcial_tomado": parcial_tomado}
    except Exception as e:
        print(f"| API ERROR | Fallo al buscar TP para ticket {ticket}: {e}")
    return {"status": "error", "tp": 0.0, "parcial_tomado": False}

@app.get("/api/export_audit_csv")
def export_audit_csv():
    """
    Exporta todos los logs de auditoría a formato CSV para análisis de datos duros.
    """
    if not firebase_inicializado or db is None:
        raise HTTPException(status_code=503, detail="Firebase no inicializado")
        
    try:
        asegurar_cache_firebase()
        
        output = StringIO()
        writer = csv.writer(output)
        # Escribir la cabecera
        writer.writerow(["Ticket", "Timestamp", "Fecha", "Activo", "Accion", "Estrategia", "Score (%)", "Precio Ejecucion", "PNL", "Ejecutada MT5", "Motivo", "Conf. Tecnicas", "Conf. Fundamentales", "Conf. Institucionales", "Detalle Setup"])
        
        if GLOBAL_AUDIT_LOGS:
            for l in GLOBAL_AUDIT_LOGS:
                import json
                tecnicas = json.dumps(l.get("confirmaciones_tecnicas", {}))
                fundamentales = json.dumps(l.get("confirmaciones_fundamentales", {}))
                institucionales = json.dumps(l.get("confirmaciones_institucionales", {}))
                
                writer.writerow([
                    l.get("ticket", ""),
                    l.get("timestamp", ""),
                    l.get("fecha", ""),
                    l.get("activo", ""),
                    l.get("accion", ""),
                    l.get("estrategia", ""),
                    l.get("score", l.get("score_confluencias", 0)),
                    l.get("precio_ejecucion", 0.0),
                    l.get("pnl", 0.0),
                    "SÍ" if l.get("ejecutada_mt5", True) else "NO",
                    l.get("motivo", "Ejecutado" if l.get("ejecutada_mt5", True) else "Desconocido"),
                    tecnicas,
                    fundamentales,
                    institucionales,
                    l.get("detalle_setup", "")
                ])
            
        output.seek(0)
        return StreamingResponse(
            output, 
            media_type="text/csv", 
            headers={"Content-Disposition": "attachment; filename=mia_audit_logs.csv"}
        )
    except Exception as e:
        print(f"| API EXPORT CSV ERROR | {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/chart_data/{symbol}")
async def get_chart_data(symbol: str):
    try:
        import yfinance as yf
        import pandas as pd
        # Mapeo de símbolos de Mia a Yahoo Finance
        mapa = {
            "EURUSD": "EURUSD=X",
            "GBPUSD": "GBPUSD=X",
            "GBPJPY": "GBPJPY=X",
            "XAUUSD": "GC=F",
            "NASDAQ100": "NQ=F",
            "BTCUSD": "BTC-USD",
            "US30": "YM=F",
            "SP500": "ES=F"
        }
        
        yf_symbol = mapa.get(symbol.upper(), symbol.upper())
        
        # Descargar data de 7 días, intervalos de 1 hora
        ticker = yf.Ticker(yf_symbol)
        df = ticker.history(period="7d", interval="1h")
        
        if df.empty:
            # Reintentar con símbolo spot si es oro o un futuro
            if yf_symbol == "GC=F":
                df = yf.Ticker("XAUUSD=X").history(period="7d", interval="1h")
            if df.empty:
                return {"status": "error", "message": f"No se encontraron datos para {yf_symbol}"}
            
        if not df.empty:
            df['EMA50'] = df['Close'].ewm(span=50, adjust=False).mean()
            df['EMA200'] = df['Close'].ewm(span=200, adjust=False).mean()
            
        candles = []
        ema50 = []
        ema200 = []
        
        for index, row in df.iterrows():
            ts = int(index.timestamp())
            candles.append({
                "time": ts,
                "open": float(row["Open"]),
                "high": float(row["High"]),
                "low": float(row["Low"]),
                "close": float(row["Close"])
            })
            if not pd.isna(row.get("EMA50")):
                ema50.append({"time": ts, "value": float(row["EMA50"])})
            if not pd.isna(row.get("EMA200")):
                ema200.append({"time": ts, "value": float(row["EMA200"])})
                
        # Consultar trades en Firebase para crear los marcadores visuales (flechas)
        markers = []
        try:
            asegurar_cache_firebase()
            global GLOBAL_AUDIT_LOGS
            
            # Filtramos el caché en RAM (Cero coste de lectura en Firebase)
            logs_filtrados = [log for log in GLOBAL_AUDIT_LOGS if log.get("activo", "").upper() == symbol.upper()]
            
            for data in logs_filtrados:
                accion = data.get("accion", "").upper()
                precio = float(data.get("precio_ejecucion", 0.0) or data.get("precio", 0.0))
                fecha_str = data.get("fecha", "") # Ej: 2026-06-30 08:30:00
                
                if accion in ["COMPRA", "VENTA"] and fecha_str:
                        # Convertir fecha a timestamp aproximado (UTC o local dependiendo de como se guardo)
                        # Como yfinance devuelve los index en UTC o timezone local, intentamos simplificar
                        from datetime import datetime
                        dt = datetime.strptime(fecha_str, "%Y-%m-%d %H:%M:%S")
                        ts_marker = int(dt.timestamp())
                        
                        is_buy = accion == "COMPRA"
                        markers.append({
                            "time": ts_marker,
                            "position": "belowBar" if is_buy else "aboveBar",
                            "color": "#00e68a" if is_buy else "#f85149",
                            "shape": "arrowUp" if is_buy else "arrowDown",
                            "text": "BUY" if is_buy else "SELL",
                            "size": 2
                        })
        except Exception as mk_err:
            print(f"| CHART MARKERS ERROR | {mk_err}")
            
        # Ordenar markers por tiempo para evitar errores en LightweightCharts
        markers = sorted(markers, key=lambda x: x["time"])
            
        return {
            "status": "success", 
            "data": candles,
            "ema50": ema50,
            "ema200": ema200,
            "markers": markers
        }
    except Exception as e:
        print(f"| CHART API ERROR | {e}")
        return {"status": "error", "message": str(e)}
